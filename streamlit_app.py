"""
Sustainability Framework Analyser - Streamlit App
Deploy to Streamlit Cloud for free public access.

Supports Claude Haiku 4.5 and OpenAI GPT-5.6 models for intelligent report
analysis against sustainability framework requirements.

Requirements are loaded from ReportingFrameworks_v1.xlsx (in the project repo).
"""

import streamlit as st

# Force light theme without needing .streamlit/config.toml
st._config.set_option("theme.base", "light")
st._config.set_option("theme.primaryColor", "#1C6B4A")
st._config.set_option("theme.backgroundColor", "#F5F1E8")
st._config.set_option("theme.secondaryBackgroundColor", "#EDE7D8")
st._config.set_option("theme.textColor", "#152018")

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import html
import json
from io import BytesIO
from collections import defaultdict

import report_drafter
from startup_compat import import_module_with_exports

_ANALYSIS_CORE_EXPORTS = (
    "ANALYST_MODELS",
    "AnalysisAuthenticationError",
    "HAIKU_MODEL",
    "LUNA_MODEL",
    "MODEL_CATALOG",
    "PRIMARY_MODEL",
    "REVIEWER_MODELS",
    "SENIOR_REVIEWER_MODELS",
    "TERRA_MODEL",
    "USER_SELECTABLE_MODELS",
    "analyze_report",
    "analyze_report_with_review_cascade",
    "estimate_usage_cost",
    "extract_pdf_pages",
    "format_report_text",
    "get_model_config",
    "model_picker_label",
)
import_module_with_exports("analysis_core", _ANALYSIS_CORE_EXPORTS)

from analysis_core import (
    ANALYST_MODELS,
    AnalysisAuthenticationError,
    HAIKU_MODEL,
    LUNA_MODEL,
    MODEL_CATALOG,
    PRIMARY_MODEL,
    REVIEWER_MODELS,
    SENIOR_REVIEWER_MODELS,
    TERRA_MODEL,
    USER_SELECTABLE_MODELS,
    analyze_report,
    analyze_report_with_review_cascade,
    estimate_usage_cost,
    extract_pdf_pages,
    format_report_text,
    get_model_config,
    model_picker_label,
)

# Page config
st.set_page_config(
    page_title="Sustainability Framework Analyser",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS — force light theme without breaking tabs, Plotly, or alert boxes
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Spectral:wght@400;500;600;700&family=Hanken+Grotesk:wght@400;500;600;700;800&family=IBM+Plex+Mono:wght@400;500;600&display=swap');
    /* ===== Global background ===== */
    .stApp {
        background-color: #F5F1E8;
        color: #152018;
    }
    .main .block-container {
        padding-top: 2rem;
        color: #152018;
    }

    /* ===== Typography — scoped to avoid Plotly / tab leaks ===== */
    h1, h2, h3, h4 {
        color: #152018 !important;
    }
    .stMarkdown, .stMarkdown p, .stMarkdown span, .stMarkdown li,
    .stText, .stCaption, .stSubheader {
        color: #152018 !important;
    }

    /* ===== Labels (checkbox, select, input, file uploader) ===== */
    .stCheckbox label, .stCheckbox label span,
    .stSelectbox label, .stTextInput label, .stTextArea label,
    .stFileUploader label, .stNumberInput label {
        color: #152018 !important;
    }

    /* ===== Tabs ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: #E9E3D3 !important;
        border-radius: 8px;
        padding: 10px 20px;
        color: #4B5A50 !important;
    }
    /* Selected tab — higher specificity so it wins */
    .stTabs [data-baseweb="tab"][aria-selected="true"],
    .stTabs [data-baseweb="tab"][aria-selected="true"] * {
        background-color: #0F3D2A !important;
        color: #FCFAF3 !important;
    }

    /* ===== Buttons ===== */
    .stButton > button {
        background-color: #E9E3D3 !important;
        color: #152018 !important;
        border: 1px solid #CFC7B2 !important;
    }
    .stButton > button:hover {
        background-color: #DDD5C2 !important;
        border-color: #B8B09A !important;
    }
    .stButton > button[kind="primary"],
    .stButton > button[data-testid="stBaseButton-primary"] {
        background-color: #1C6B4A !important;
        color: #FCFAF3 !important;
        border: none !important;
    }

    /* ===== Inputs (text, password, number, textarea) ===== */
    .stTextInput input, .stNumberInput input, .stTextArea textarea {
        background-color: #FCFAF3 !important;
        color: #152018 !important;
        border: 1px solid #CFC7B2 !important;
    }
    .stTextInput > div > div, .stNumberInput > div > div {
        background-color: #FCFAF3 !important;
    }
    .stNumberInput button {
        background-color: #E9E3D3 !important;
        color: #152018 !important;
        border-color: #CFC7B2 !important;
    }

    /* ===== Select boxes (dropdowns) ===== */
    [data-baseweb="select"] {
        background-color: #FCFAF3 !important;
    }
    [data-baseweb="select"] > div {
        background-color: #FCFAF3 !important;
        border-color: #CFC7B2 !important;
    }
    [data-baseweb="select"] span, [data-baseweb="select"] div {
        color: #152018 !important;
    }
    /* Dropdown menu */
    [data-baseweb="popover"], [data-baseweb="menu"] {
        background-color: #FCFAF3 !important;
    }
    [data-baseweb="popover"] li, [data-baseweb="menu"] li {
        color: #152018 !important;
    }

    /* ===== File uploader ===== */
    .stFileUploader section {
        background-color: #EDE7D8 !important;
        border-color: #CFC7B2 !important;
    }
    .stFileUploader section span, .stFileUploader section small,
    .stFileUploader section div {
        color: #4B5A50 !important;
    }
    .stFileUploader section button {
        background-color: #FCFAF3 !important;
        color: #152018 !important;
        border: 1px solid #CFC7B2 !important;
    }
    [data-testid="stFileUploaderDropzone"] {
        background-color: #EDE7D8 !important;
        border-color: #CFC7B2 !important;
    }
    [data-testid="stFileUploaderDropzone"] * {
        color: #4B5A50 !important;
    }
    /* Uploaded file name */
    [data-testid="stFileUploaderFile"] span,
    [data-testid="stFileUploaderFile"] div {
        color: #152018 !important;
    }

    /* ===== Expanders (Results) ===== */
    [data-testid="stExpander"] {
        background-color: #FCFAF3 !important;
        border-color: #DDD5C2 !important;
    }
    [data-testid="stExpander"] summary,
    [data-testid="stExpander"] summary span,
    [data-testid="stExpander"] summary p {
        color: #152018 !important;
    }

    /* ===== Alert boxes — inherit their own colours ===== */
    .stAlert, .stAlert p, .stAlert span {
        color: inherit !important;
    }

    /* ===== Plotly — do NOT override; let it manage its own text ===== */
    .js-plotly-plot, .js-plotly-plot * {
        /* no color override */
    }

    /* ===== Badge classes ===== */
    .framework-card {
        background-color: #EDE7D8;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .badge-covers {
        background-color: #E8F2EA; color: #1C6B4A !important;
        padding: 4px 12px; border-radius: 12px; font-weight: 600; font-size: 13px;
    }
    .badge-partly {
        background-color: #FBF0D8; color: #B07A18 !important;
        padding: 4px 12px; border-radius: 12px; font-weight: 600; font-size: 13px;
    }
    .badge-doesnt {
        background-color: #F8E3DD; color: #B4472F !important;
        padding: 4px 12px; border-radius: 12px; font-weight: 600; font-size: 13px;
    }
    /* ===== Dark panels keep light text ===== */
    .stMarkdown .terra-dark-card, .stMarkdown .terra-dark-card p,
    .stMarkdown .terra-dark-card span, .stMarkdown .terra-dark-card div,
    .stMarkdown .terra-dark-card strong,
    .terra-dark-card p, .terra-dark-card span, .terra-dark-card strong {
        color: #E8EFE8 !important;
    }
    .stMarkdown .terra-dark-card .t-strong,
    .terra-dark-card .t-strong { color: #FCFAF3 !important; }
    .stMarkdown .terra-dark-card .t-soft,
    .terra-dark-card .t-soft { color: #9FBAA8 !important; }

    /* ===== Terra typography & polish ===== */
    html, body, .stApp, .stMarkdown, .stButton > button, input, textarea, [data-baseweb="select"] span {
        font-family: 'Hanken Grotesk', sans-serif;
    }
    h1, h2, h3 {
        font-family: 'Spectral', serif !important;
        font-weight: 600 !important;
        letter-spacing: -0.01em;
    }
    code, pre { font-family: 'IBM Plex Mono', monospace; }
    .stButton > button { border-radius: 9px !important; font-weight: 600 !important; }
    .stButton > button[kind="primary"],
    .stButton > button[data-testid="stBaseButton-primary"] {
        box-shadow: 0 8px 18px -8px rgba(28,107,74,.55) !important;
    }
    .stTabs [data-baseweb="tab"] { border-radius: 9px; font-weight: 600; }
    [data-testid="stExpander"] { border-radius: 12px !important; }
</style>
""", unsafe_allow_html=True)

# ============================================
# DATA
# ============================================

FRAMEWORK_COLORS = {
    "TCFD": "#3b82f6",
    "TNFD": "#10b981",
    "PRA": "#f59e0b",
    "IFRS S1": "#ef4444",
    "IFRS S2": "#dc2626",
    "TPT": "#8b5cf6",
    "BMA": "#ec4899",
    "MAS": "#14b8a6",
    "ESRS E1": "#f97316",
    "ESRS E4": "#fb923c",
    "OSFI": "#06b6d4",
    "SBTi": "#a855f7",
    "PSI": "#64748b"
}

FRAMEWORK_FULL_NAMES = {
    "TCFD": "Task Force on Climate-related Financial Disclosures",
    "TNFD": "Taskforce on Nature-related Financial Disclosures",
    "PRA": "Prudential Regulation Authority",
    "IFRS S1": "IFRS S1 – General Requirements for Disclosure of Sustainability-related Financial Information",
    "IFRS S2": "IFRS S2 – Climate-related Disclosures",
    "TPT": "Transition Plan Taskforce",
    "BMA": "Bermuda Monetary Authority",
    "MAS": "Monetary Authority of Singapore",
    "ESRS E1": "European Sustainability Reporting Standards – Climate Change (E1)",
    "ESRS E4": "European Sustainability Reporting Standards – Biodiversity and Ecosystems (E4)",
    "OSFI": "Office of the Superintendent of Financial Institutions",
    "SBTi": "Science Based Targets initiative",
    "PSI": "Principles for Sustainable Insurance"
}

FRAMEWORK_URLS = {
    "TCFD": "https://www.fsb-tcfd.org/",
    "TNFD": "https://tnfd.global/",
    "PRA": "https://www.bankofengland.co.uk/prudential-regulation/publication/2019/enhancing-banks-and-insurers-approaches-to-managing-the-financial-risks-from-climate-change-ss",
    "IFRS S1": "https://www.ifrs.org/issued-standards/ifrs-sustainability-standards-navigator/ifrs-s1-general-requirements/",
    "IFRS S2": "https://www.ifrs.org/issued-standards/ifrs-sustainability-standards-navigator/ifrs-s2-climate-related-disclosures/",
    "TPT": "https://transitiontaskforce.net/",
    "BMA": "https://www.bma.bm/",
    "MAS": "https://www.mas.gov.sg/regulation/guidelines/guidelines-on-environmental-risk-management",
    "ESRS E1": "https://www.efrag.org/en/sustainability-reporting/esrs-workstreams/esrs-e1-climate-change",
    "ESRS E4": "https://www.efrag.org/en/sustainability-reporting/esrs-workstreams/esrs-e4-biodiversity-and-ecosystems",
    "OSFI": "https://www.osfi-bsif.gc.ca/en/guidance/guidance-library/climate-risk-management",
    "SBTi": "https://sciencebasedtargets.org/",
    "PSI": "https://www.unepfi.org/insurance/insurance/",
}

ADOPTION_DICT = {
    "TCFD": ["Canada", "France", "Germany", "Italy", "Japan", "United Kingdom", "USA", "New Zealand", "Switzerland", "Singapore", "Brazil", "China", "South Africa"],
    "TNFD": ["Brazil", "China", "Colombia", "Costa Rica", "Egypt", "India", "Indonesia", "Kenya", "Malaysia", "Mexico", "Morocco", "Nigeria", "Peru", "Philippines", "South Africa"],
    "PRA": ["United Kingdom"],
    "IFRS S1": ["Turkey", "Bangladesh", "Brazil", "Australia", "Japan", "United Kingdom", "Canada", "Singapore", "New Zealand", "Nigeria", "South Africa", "Malaysia", "China"],
    "IFRS S2": ["Turkey", "Bangladesh", "Brazil", "Australia", "Japan", "United Kingdom", "Canada", "Singapore", "New Zealand", "Nigeria", "South Africa", "Malaysia", "China"],
    "TPT": ["United Kingdom"],
    "BMA": ["Bermuda"],
    "MAS": ["Singapore"],
    "ESRS E1": ["Austria", "Belgium", "Bulgaria", "Croatia", "Cyprus", "Czech Republic", "Denmark", "Estonia", "Finland", "France", "Germany", "Greece", "Hungary", "Ireland", "Italy", "Latvia", "Lithuania", "Luxembourg", "Malta", "Netherlands", "Poland", "Portugal", "Romania", "Slovakia", "Slovenia", "Spain", "Sweden"],
    "ESRS E4": ["Austria", "Belgium", "Bulgaria", "Croatia", "Cyprus", "Czech Republic", "Denmark", "Estonia", "Finland", "France", "Germany", "Greece", "Hungary", "Ireland", "Italy", "Latvia", "Lithuania", "Luxembourg", "Malta", "Netherlands", "Poland", "Portugal", "Romania", "Slovakia", "Slovenia", "Spain", "Sweden"],
    "OSFI": ["Canada"],
    "SBTi": ["Japan", "United Kingdom", "USA", "China", "Germany", "France", "India", "Italy", "Canada", "South Korea", "Mexico", "Brazil", "Australia", "South Africa", "Turkey", "Romania", "Malta"],
    "PSI": ["Japan", "United Kingdom", "USA", "Germany", "France", "Brazil", "Australia", "South Africa", "China", "India", "Singapore", "Canada", "Switzerland", "Netherlands", "Sweden"]
}

COUNTRY_COORDS = {
    "Canada": {"lat": 56.13, "lon": -106.35},
    "USA": {"lat": 37.09, "lon": -95.71},
    "Mexico": {"lat": 23.63, "lon": -102.55},
    "Brazil": {"lat": -14.24, "lon": -51.93},
    "Colombia": {"lat": 4.57, "lon": -74.30},
    "Costa Rica": {"lat": 9.75, "lon": -83.75},
    "Peru": {"lat": -9.19, "lon": -75.02},
    "United Kingdom": {"lat": 55.38, "lon": -3.44},
    "France": {"lat": 46.23, "lon": 2.21},
    "Germany": {"lat": 51.17, "lon": 10.45},
    "Italy": {"lat": 41.87, "lon": 12.57},
    "Spain": {"lat": 40.46, "lon": -3.75},
    "Switzerland": {"lat": 46.82, "lon": 8.23},
    "Austria": {"lat": 47.52, "lon": 14.55},
    "Belgium": {"lat": 50.50, "lon": 4.47},
    "Netherlands": {"lat": 52.13, "lon": 5.29},
    "Poland": {"lat": 51.92, "lon": 19.15},
    "Sweden": {"lat": 60.13, "lon": 18.64},
    "Denmark": {"lat": 56.26, "lon": 9.50},
    "Finland": {"lat": 61.92, "lon": 25.75},
    "Greece": {"lat": 39.07, "lon": 21.82},
    "Portugal": {"lat": 39.40, "lon": -8.22},
    "Ireland": {"lat": 53.14, "lon": -7.69},
    "Bulgaria": {"lat": 42.73, "lon": 25.49},
    "Romania": {"lat": 45.94, "lon": 24.97},
    "Hungary": {"lat": 47.16, "lon": 19.50},
    "Czech Republic": {"lat": 49.82, "lon": 15.47},
    "Slovakia": {"lat": 48.67, "lon": 19.70},
    "Slovenia": {"lat": 46.15, "lon": 14.99},
    "Croatia": {"lat": 45.10, "lon": 15.20},
    "Estonia": {"lat": 58.60, "lon": 25.01},
    "Latvia": {"lat": 56.88, "lon": 24.60},
    "Lithuania": {"lat": 55.17, "lon": 23.88},
    "Cyprus": {"lat": 35.13, "lon": 33.43},
    "Malta": {"lat": 35.94, "lon": 14.38},
    "Luxembourg": {"lat": 49.82, "lon": 6.13},
    "Turkey": {"lat": 38.96, "lon": 35.24},
    "Egypt": {"lat": 26.82, "lon": 30.80},
    "Morocco": {"lat": 31.79, "lon": -7.09},
    "South Africa": {"lat": -30.56, "lon": 22.94},
    "Nigeria": {"lat": 9.08, "lon": 8.68},
    "Kenya": {"lat": -0.02, "lon": 37.91},
    "Japan": {"lat": 36.20, "lon": 138.25},
    "South Korea": {"lat": 35.91, "lon": 127.77},
    "China": {"lat": 35.86, "lon": 104.20},
    "India": {"lat": 20.59, "lon": 78.96},
    "Singapore": {"lat": 1.35, "lon": 103.82},
    "Malaysia": {"lat": 4.21, "lon": 101.98},
    "Indonesia": {"lat": -0.79, "lon": 113.92},
    "Philippines": {"lat": 12.88, "lon": 121.77},
    "Bangladesh": {"lat": 23.68, "lon": 90.36},
    "Australia": {"lat": -25.27, "lon": 133.78},
    "New Zealand": {"lat": -40.90, "lon": 174.89},
    "Bermuda": {"lat": 32.32, "lon": -64.76}
}

# Similarity data from the notebook
# NOTE: Similarity keys use the original framework names (TCFD, TNFD, PRA, IFRS, TPT, BMA, MAS, ESRS, OSFI, SBTi).
# For split frameworks (IFRS S1/S2, ESRS E1/E4), similarity lookups fall back to the parent name.
SIMILARITY_PARENT_MAP = {
    "IFRS S1": "IFRS", "IFRS S2": "IFRS",
    "ESRS E1": "ESRS", "ESRS E4": "ESRS",
}


# ============================================
# LOAD SIMILARITY DATA FROM CSV FILES
# ============================================

SIMILARITY_METRIC_TYPES = [
    "all_metrics", "governance", "strategy", "risk", "metrics", "disclosure"
]

@st.cache_data
def load_similarity_data():
    """
    Load similarity matrices from CSV files in the same directory as the app.
    Expects files named: similarity_network_table_{metric_type}.csv
    Each CSV should have columns: Framework 1, Framework 2, Similarity
    Returns a dict mapping metric_type -> DataFrame.
    """
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    data = {}

    for metric in SIMILARITY_METRIC_TYPES:
        filename = f"similarity_network_table_{metric}.csv"
        filepath = os.path.join(base_dir, filename)
        try:
            df = pd.read_csv(filepath)
            # Ensure expected columns exist
            if all(
                col in df.columns
                for col in ["Framework 1", "Framework 2", "Similarity"]
            ):
                data[metric] = df
            else:
                st.warning(
                    f"Similarity file {filename} missing expected columns. "
                    f"Found: {list(df.columns)}"
                )
        except FileNotFoundError:
            pass  # Metric type won't be available
        except Exception as e:
            st.warning(f"Could not load {filename}: {e}")

    return data



# ============================================
# LOAD FRAMEWORK REQUIREMENTS FROM EXCEL
# ============================================

@st.cache_data
def load_framework_requirements():
    """
    Load framework requirements from ReportingFrameworks_v1.xlsx.
    Returns a tuple:
      - requirements: { framework: { topic: [recommendation_1, ...] } }
      - references:   { (framework, recommendation): reference_string }
    Recommendations are deduplicated per framework+topic.
    """
    import os

    # Try multiple paths: same directory as script, then common locations
    possible_paths = [
        os.path.join(os.path.dirname(__file__), "ReportingFrameworks_v1.xlsx"),
        "ReportingFrameworks_v1.xlsx",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "ReportingFrameworks_v1.xlsx"),
    ]

    df = None
    for path in possible_paths:
        if os.path.exists(path):
            df = pd.read_excel(path, engine="openpyxl")
            break

    if df is None:
        st.error(
            "Could not find ReportingFrameworks_v1.xlsx. "
            "Please ensure the file is in the same directory as this script."
        )
        return {}, {}

    requirements = defaultdict(lambda: defaultdict(list))
    references = {}

    for _, row in df.iterrows():
        framework = row.get("Framework")
        topic = row.get("Topic")
        recommendation = row.get("Recommendation")
        reference = row.get("Reference")

        if pd.isna(framework) or pd.isna(topic) or pd.isna(recommendation):
            continue

        framework = str(framework).strip()
        topic = str(topic).strip()
        recommendation = str(recommendation).strip()

        # Deduplicate
        if recommendation not in requirements[framework][topic]:
            requirements[framework][topic].append(recommendation)

        # Store source reference (e.g. "IFRS S2 6", "TCFD Governance A")
        if not pd.isna(reference):
            references[(framework, recommendation)] = str(reference).strip()

    # Convert defaultdicts to regular dicts for caching
    return (
        {fw: dict(topics) for fw, topics in requirements.items()},
        references,
    )


# ============================================
# CLASSIFICATION HELPERS
# ============================================

CLASSIFICATION_COVERS = "Covers the framework"
CLASSIFICATION_PARTLY = "Partly covers the framework"
CLASSIFICATION_DOESNT = "Doesn't cover the framework"

ALL_CLASSIFICATIONS = [CLASSIFICATION_COVERS, CLASSIFICATION_PARTLY, CLASSIFICATION_DOESNT]

CLASSIFICATION_COLORS = {
    CLASSIFICATION_COVERS: "#1C6B4A",
    CLASSIFICATION_PARTLY: "#C98A2B",
    CLASSIFICATION_DOESNT: "#B4472F",
}

CLASSIFICATION_BADGES = {
    CLASSIFICATION_COVERS: "badge-covers",
    CLASSIFICATION_PARTLY: "badge-partly",
    CLASSIFICATION_DOESNT: "badge-doesnt",
}


def classification_to_score(classification):
    """Map classification to a numeric value for summary statistics."""
    if classification == CLASSIFICATION_COVERS:
        return 1.0
    elif classification == CLASSIFICATION_PARTLY:
        return 0.5
    else:
        return 0.0


# ============================================
# HELPER FUNCTIONS
# ============================================

def extract_text_from_pdf(pdf_file):
    """Backward-compatible text-only PDF extraction for other app tabs."""
    return [page["text"] for page in extract_pdf_pages(pdf_file)]


def run_model_analysis(
    report_text,
    selected_frameworks,
    api_key,
    framework_requirements,
    progress_bar=None,
    requirement_refs=None,
    report_pages=None,
    use_batch=True,
    existing_batch_id=None,
    track_pending_batch=False,
    model_id=PRIMARY_MODEL,
):
    """Run the confidence-aware multimodal analysis pipeline."""

    def update_progress(value):
        if progress_bar:
            progress_bar.progress(value)

    def show_status(level, message):
        getattr(st, level, st.info)(message)

    def remember_batch_id(batch_id):
        pending = st.session_state.get("pending_analysis")
        if pending is not None:
            pending["batch_id"] = batch_id

    return analyze_report(
        report_text=report_text,
        selected_frameworks=selected_frameworks,
        api_key=api_key,
        framework_requirements=framework_requirements,
        framework_full_names=FRAMEWORK_FULL_NAMES,
        requirement_refs=requirement_refs,
        report_pages=report_pages,
        use_batch=use_batch,
        existing_batch_id=existing_batch_id,
        batch_id_callback=(remember_batch_id if track_pending_batch else None),
        progress_callback=update_progress,
        status_callback=show_status,
        model_id=model_id,
    )


def run_review_cascade(
    report_text,
    selected_frameworks,
    anthropic_api_key,
    openai_api_key,
    framework_requirements,
    progress_bar=None,
    requirement_refs=None,
    report_pages=None,
    analyst_model_id=HAIKU_MODEL,
    reviewer_model_id=LUNA_MODEL,
    senior_reviewer_model_id=TERRA_MODEL,
):
    """Run the selected three-role review cascade using standard calls."""

    def update_progress(value):
        if progress_bar:
            progress_bar.progress(value)

    def show_status(level, message):
        getattr(st, level, st.info)(message)

    return analyze_report_with_review_cascade(
        report_text=report_text,
        selected_frameworks=selected_frameworks,
        anthropic_api_key=anthropic_api_key,
        openai_api_key=openai_api_key,
        framework_requirements=framework_requirements,
        framework_full_names=FRAMEWORK_FULL_NAMES,
        requirement_refs=requirement_refs,
        report_pages=report_pages,
        progress_callback=update_progress,
        status_callback=show_status,
        analyst_model_id=analyst_model_id,
        reviewer_model_id=reviewer_model_id,
        senior_reviewer_model_id=senior_reviewer_model_id,
    )


def render_model_api_key(model_id, widget_key):
    """Render the credential field belonging to the selected provider."""
    model = get_model_config(model_id)
    provider = "Anthropic" if model["provider"] == "anthropic" else "OpenAI"
    secret_name = model["secret_name"]
    secrets_key = st.secrets.get(secret_name, "")
    if secrets_key:
        st.markdown(
            '<div style="background:#E8F2EA;border:1px solid #C6E0CC;'
            'border-radius:8px;padding:10px;font-size:13px;color:#1C6B4A;">'
            f'API key configured for {provider}</div>',
            unsafe_allow_html=True,
        )
        return secrets_key

    placeholder = "sk-ant-..." if provider == "Anthropic" else "sk-..."
    console = (
        "console.anthropic.com"
        if provider == "Anthropic"
        else "platform.openai.com/api-keys"
    )
    return st.text_input(
        f"{provider} API Key",
        type="password",
        placeholder=placeholder,
        help=(
            f"Required for {model['label']}. Your key is not stored. "
            f"Get one at {console}"
        ),
        key=widget_key,
    )


def render_model_price_caption(model_id):
    """Show standard, cached, and batch list prices next to the picker."""
    model = get_model_config(model_id)
    long_context_note = (
        " Inputs above 272K tokens use higher long-context rates."
        if model.get("long_context_threshold")
        else ""
    )
    st.caption(
        f"{model['description']}. USD per 1M tokens — standard: "
        f"${model['input_price']:g} input / ${model['output_price']:g} "
        f"output; cache read / write: ${model['cached_input_price']:g} / "
        f"${model['cache_write_price']:g}; Batch API (cheaper, slower): "
        f"${model['batch_input_price']:g} input / "
        f"${model['batch_output_price']:g} output. Vision and reasoning "
        f"change token usage.{long_context_note}"
    )


ANALYSIS_STRATEGY_SINGLE = "Single model"
ANALYSIS_STRATEGY_CASCADE = "Reviewed cascade"

CASCADE_STATUS_LABELS = {
    "analyst_reviewer_agree": "Analyst + reviewer agree",
    "senior_reviewer_adjudicated": "Senior reviewer adjudicated",
    "reviewer_failed": "Reviewer incomplete",
    "senior_reviewer_failed": "Senior reviewer incomplete",
    # Retain labels for results created before configurable cascade roles.
    "haiku_luna_agree": "Haiku + Luna agree",
    "terra_adjudicated": "Terra adjudicated",
    "three_way_disagreement": "Three-way disagreement",
    "luna_review_failed": "Luna review incomplete",
    "terra_review_failed": "Terra review incomplete",
}
CASCADE_PROVISIONAL_STATUSES = {
    "three_way_disagreement",
    "reviewer_failed",
    "senior_reviewer_failed",
    "luna_review_failed",
    "terra_review_failed",
}
LEGACY_CASCADE_ROLE_KEYS = {
    "analyst": "haiku",
    "reviewer": "luna",
    "senior_reviewer": "terra",
}


def get_cascade_vote(result, role):
    """Return one role's saved cascade verdict, including legacy runs."""
    votes = result.get("model_verdicts", {})
    if not isinstance(votes, dict):
        return {}
    vote = votes.get(role)
    if vote is None:
        vote = votes.get(LEGACY_CASCADE_ROLE_KEYS.get(role, ""), {})
    return vote if isinstance(vote, dict) else {}


def cascade_role_model(result, role):
    """Return the model ID assigned to a cascade role."""
    role_models = result.get("role_models", {})
    if isinstance(role_models, dict) and role_models.get(role):
        return role_models[role]
    legacy_models = {
        "analyst": HAIKU_MODEL,
        "reviewer": LUNA_MODEL,
        "senior_reviewer": TERRA_MODEL,
    }
    return legacy_models[role]


def cascade_role_label(result, role):
    model_id = cascade_role_model(result, role)
    try:
        return get_model_config(model_id)["label"]
    except ValueError:
        return str(model_id)


def is_review_cascade_result(result):
    return result.get("analysis_mode") == "review_cascade"


def is_provisional_cascade_result(result):
    return (
        is_review_cascade_result(result)
        and result.get("cascade_status") in CASCADE_PROVISIONAL_STATUSES
    )


def result_needs_human_review(result):
    """Mirror the UI review queue for exports and summary counts."""
    return bool(
        result.get("needs_human_review")
        or result.get("confidence", "low") == "low"
    )


def cascade_status_label(result):
    status = result.get("cascade_status", "")
    analyst = cascade_role_label(result, "analyst")
    reviewer = cascade_role_label(result, "reviewer")
    senior = cascade_role_label(result, "senior_reviewer")
    dynamic_labels = {
        "analyst_reviewer_agree": f"{analyst} + {reviewer} agree",
        "senior_reviewer_adjudicated": f"{senior} adjudicated",
        "reviewer_failed": f"{reviewer} review incomplete",
        "senior_reviewer_failed": f"{senior} review incomplete",
    }
    if status in dynamic_labels:
        return dynamic_labels[status]
    return CASCADE_STATUS_LABELS.get(status, str(status).replace("_", " ").title())


def build_cascade_review_html(result):
    """Return escaped badge and compact audit trail for a cascade result."""
    if not is_review_cascade_result(result):
        return "", ""

    status = result.get("cascade_status", "")
    status_label = html.escape(cascade_status_label(result))
    status_colors = {
        "analyst_reviewer_agree": ("#E8F2EA", "#1C6B4A"),
        "senior_reviewer_adjudicated": ("#FBF0D8", "#977322"),
        "reviewer_failed": ("#F8E3DD", "#B4472F"),
        "senior_reviewer_failed": ("#F8E3DD", "#B4472F"),
        "haiku_luna_agree": ("#E8F2EA", "#1C6B4A"),
        "terra_adjudicated": ("#FBF0D8", "#977322"),
        "three_way_disagreement": ("#F8E3DD", "#B4472F"),
        "luna_review_failed": ("#F8E3DD", "#B4472F"),
        "terra_review_failed": ("#F8E3DD", "#B4472F"),
    }
    status_bg, status_fg = status_colors.get(
        status, ("#EDE7D8", "#4B5A50")
    )
    badge = (
        f'<span style="white-space:nowrap;background:{status_bg};'
        f'color:{status_fg};padding:3px 8px;border-radius:10px;'
        f'font-size:10px;font-weight:700;">{status_label}</span>'
    )

    vote_rows = []
    for role, role_title in (
        ("analyst", "Analyst"),
        ("reviewer", "Reviewer"),
        ("senior_reviewer", "Senior reviewer"),
    ):
        display_name = (
            f"{role_title} — {cascade_role_label(result, role)}"
        )
        vote = get_cascade_vote(result, role)
        if not vote:
            if role == "senior_reviewer" and status in {
                "analyst_reviewer_agree",
                "haiku_luna_agree",
            }:
                vote_rows.append(
                    '<div style="font-size:11px;color:#6E796F;">'
                    f"<strong>{html.escape(display_name)}:</strong> Not "
                    "needed — analyst and reviewer agreed.</div>"
                )
            elif role == "reviewer" and status in {
                "reviewer_failed",
                "luna_review_failed",
            }:
                vote_rows.append(
                    '<div style="font-size:11px;color:#B4472F;">'
                    f"<strong>{html.escape(display_name)}:</strong> Review did "
                    "not complete.</div>"
                )
            elif role == "senior_reviewer" and status in {
                "senior_reviewer_failed",
                "terra_review_failed",
            }:
                vote_rows.append(
                    '<div style="font-size:11px;color:#B4472F;">'
                    f"<strong>{html.escape(display_name)}:</strong> "
                    "Adjudication did not complete."
                    "</div>"
                )
            continue
        verdict = html.escape(str(vote.get("classification") or ""))
        confidence = html.escape(str(vote.get("confidence") or ""))
        confidence_reason = html.escape(
            str(vote.get("confidence_reason") or "")
        )
        rationale = html.escape(str(vote.get("rationale") or ""))
        extracts = vote.get("relevant_extracts", [])
        if not isinstance(extracts, list):
            extracts = []
        evidence = "<br>".join(
            html.escape(str(extract)) for extract in extracts
        )
        vote_rows.append(
            '<div style="font-size:11px;color:#3B4A40;margin-top:5px;">'
            f"<strong>{html.escape(display_name)}:</strong> {verdict}"
            f"{f' · {confidence} confidence' if confidence else ''}"
            f"{f'<br><strong>Confidence reason:</strong> {confidence_reason}' if confidence_reason else ''}"
            f"{f'<br><strong>Rationale:</strong> {rationale}' if rationale else ''}"
            f"{f'<br><strong>Evidence:</strong><br>{evidence}' if evidence else ''}"
            "</div>"
        )

    review_flag = (
        '<div style="font-size:11px;color:#B4472F;margin-top:6px;'
        'font-weight:700;">Human review required</div>'
        if result_needs_human_review(result)
        else ""
    )
    section = (
        '<div style="background:#FCFAF3;border:1px solid #DDD5C2;'
        'border-radius:6px;padding:9px 10px;margin-top:9px;">'
        f'<div style="font-size:11px;color:#152018;font-weight:700;">'
        f"Review trail — {status_label}</div>"
        f"{''.join(vote_rows)}{review_flag}</div>"
    )
    return badge, section


def get_explanation(score):
    if score >= 0.5:
        return "Strong alignment - document comprehensively addresses this requirement"
    elif score >= 0.35:
        return "Good alignment - document covers key aspects of this requirement"
    elif score >= 0.25:
        return "Partial alignment - document touches on some aspects but could be more comprehensive"
    elif score >= 0.15:
        return "Weak alignment - limited coverage of this requirement"
    else:
        return "Minimal alignment - requirement not substantially addressed in document"


def get_score_color(score):
    if score >= 0.4:
        return "score-high"
    elif score >= 0.3:
        return "score-medium"
    elif score >= 0.2:
        return "score-low"
    else:
        return "score-verylow"


def get_similarity_for_framework(df, framework):
    """Get similarity scores for a specific framework.
    For split frameworks (e.g. IFRS S1), falls back to the parent name (IFRS) in similarity data.
    """
    lookup_name = SIMILARITY_PARENT_MAP.get(framework, framework)

    mask = (df['Framework 1'] == lookup_name) | (df['Framework 2'] == lookup_name)
    filtered = df[mask].copy()

    result = []
    for _, row in filtered.iterrows():
        other = row['Framework 2'] if row['Framework 1'] == lookup_name else row['Framework 1']
        result.append({
            'framework': other,
            'similarity': row['Similarity']
        })

    return sorted(result, key=lambda x: x['similarity'], reverse=True)


def generate_results_excel(results, framework_summaries):
    """Generate a formatted Excel workbook from analysis results."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F3D2A")
    green_fill = PatternFill("solid", fgColor="E8F2EA")
    amber_fill = PatternFill("solid", fgColor="FBF0D8")
    red_fill = PatternFill("solid", fgColor="F8E3DD")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def safe_excel_value(value):
        """Prevent report/model text from becoming an executable formula."""
        if isinstance(value, str):
            stripped = value.lstrip()
            if (
                stripped.startswith(("=", "+", "-", "@"))
                or value.startswith(("\t", "\r", "\n"))
            ):
                return "'" + value
        return value

    summary_headers = [
        "Framework", "Covers", "Partly Covers", "Doesn't Cover",
        "Total Requirements", "Low-confidence Review",
        "Analyst + Reviewer Agreements", "Senior Adjudications",
        "Three-way Disagreements", "Incomplete Cascade Reviews",
        "Human Review",
    ]
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = 2
    for fw, s in framework_summaries.items():
        counts = s.get("counts", {})
        ws_summary.cell(row=row, column=1, value=fw).border = thin_border
        c_cell = ws_summary.cell(row=row, column=2, value=counts.get(CLASSIFICATION_COVERS, 0))
        c_cell.fill = green_fill
        c_cell.border = thin_border
        p_cell = ws_summary.cell(row=row, column=3, value=counts.get(CLASSIFICATION_PARTLY, 0))
        p_cell.fill = amber_fill
        p_cell.border = thin_border
        d_cell = ws_summary.cell(row=row, column=4, value=counts.get(CLASSIFICATION_DOESNT, 0))
        d_cell.fill = red_fill
        d_cell.border = thin_border
        ws_summary.cell(row=row, column=5, value=s.get("total", 0)).border = thin_border
        ws_summary.cell(
            row=row, column=6, value=s.get("low_confidence", 0)
        ).border = thin_border
        framework_results = [
            result for result in results if result.get("framework") == fw
        ]
        agreement_count = sum(
            result.get("cascade_status") in {
                "analyst_reviewer_agree",
                "haiku_luna_agree",
            }
            for result in framework_results
        )
        terra_count = sum(
            result.get("cascade_status") in {
                "senior_reviewer_adjudicated",
                "terra_adjudicated",
            }
            for result in framework_results
        )
        three_way_count = sum(
            result.get("cascade_status") == "three_way_disagreement"
            for result in framework_results
        )
        incomplete_count = sum(
            result.get("cascade_status") in {
                "reviewer_failed",
                "senior_reviewer_failed",
                "luna_review_failed",
                "terra_review_failed",
            }
            for result in framework_results
        )
        human_review_count = sum(
            result_needs_human_review(result)
            for result in framework_results
        )
        for column, value in enumerate(
            (
                agreement_count,
                terra_count,
                three_way_count,
                incomplete_count,
                human_review_count,
            ),
            start=7,
        ):
            ws_summary.cell(row=row, column=column, value=value).border = (
                thin_border
            )
        row += 1

    for col_letter in [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"
    ]:
        ws_summary.column_dimensions[col_letter].width = 22

    # --- Sheet 2: Detailed Results ---
    ws_detail = wb.create_sheet("Detailed Results")
    detail_headers = [
        "Framework", "Topic", "Reference", "Requirement", "Classification",
        "Cascade Status", "Human Review Required",
        "Analyst Model and Verdict", "Analyst Confidence",
        "Analyst Review Detail",
        "Reviewer Model and Verdict", "Reviewer Confidence",
        "Reviewer Review Detail",
        "Senior Reviewer Model and Verdict", "Senior Reviewer Confidence",
        "Senior Reviewer Review Detail",
        "Confidence", "Confidence Reason", "Rationale", "Relevant Extracts",
    ]
    for col, h in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    def write_result_row(worksheet, row_number, result):
        """Write a final result plus optional cascade audit trail."""
        def vote_review_detail(vote):
            if not vote:
                return ""
            detail_lines = []
            confidence_reason = vote.get("confidence_reason", "")
            if confidence_reason:
                detail_lines.append(
                    f"Confidence reason: {confidence_reason}"
                )
            rationale = vote.get("rationale", "")
            if rationale:
                detail_lines.append(f"Rationale: {rationale}")
            extracts = vote.get("relevant_extracts", [])
            if isinstance(extracts, list) and extracts:
                detail_lines.append(
                    "Evidence: " + "; ".join(str(item) for item in extracts)
                )
            return "\n".join(detail_lines)

        analyst_vote = get_cascade_vote(result, "analyst")
        reviewer_vote = get_cascade_vote(result, "reviewer")
        senior_vote = get_cascade_vote(result, "senior_reviewer")
        is_cascade = is_review_cascade_result(result)
        status = result.get("cascade_status", "") if is_cascade else ""
        human_review = result_needs_human_review(result)
        values = [
            result["framework"],
            prettify_topic_name(result["topic"]),
            result.get("reference", ""),
            result["requirement"],
            result["classification"],
            cascade_status_label(result) if is_cascade else "",
            "Yes" if human_review else "No",
            (
                f"{cascade_role_label(result, 'analyst')}: "
                f"{analyst_vote.get('classification', '')}"
                if analyst_vote else ""
            ),
            analyst_vote.get("confidence", ""),
            vote_review_detail(analyst_vote),
            (
                f"{cascade_role_label(result, 'reviewer')}: "
                f"{reviewer_vote.get('classification', '')}"
                if reviewer_vote else ""
            ),
            reviewer_vote.get("confidence", ""),
            vote_review_detail(reviewer_vote),
            (
                f"{cascade_role_label(result, 'senior_reviewer')}: "
                f"{senior_vote.get('classification', '')}"
                if senior_vote else ""
            ),
            senior_vote.get("confidence", ""),
            vote_review_detail(senior_vote),
            result.get("confidence", "low").title(),
            result.get("confidence_reason", ""),
            result.get("rationale", ""),
            "; ".join(result.get("relevant_extracts", [])),
        ]
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(
                row=row_number,
                column=column,
                value=safe_excel_value(value),
            )
            cell.border = thin_border
            if column in {4, 10, 13, 16, 18, 19, 20}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        class_cell = worksheet.cell(row=row_number, column=5)
        if result["classification"] == CLASSIFICATION_COVERS:
            class_cell.fill = green_fill
        elif result["classification"] == CLASSIFICATION_PARTLY:
            class_cell.fill = amber_fill
        else:
            class_cell.fill = red_fill

        status_cell = worksheet.cell(row=row_number, column=6)
        if status in {"analyst_reviewer_agree", "haiku_luna_agree"}:
            status_cell.fill = green_fill
        elif status in {
            "senior_reviewer_adjudicated",
            "terra_adjudicated",
        }:
            status_cell.fill = amber_fill
        elif status == "three_way_disagreement":
            status_cell.fill = red_fill

        human_review_cell = worksheet.cell(row=row_number, column=7)
        human_review_cell.fill = red_fill if human_review else green_fill

        confidence_cell = worksheet.cell(row=row_number, column=17)
        if result.get("confidence") == "low":
            confidence_cell.fill = red_fill
        elif result.get("confidence") == "medium":
            confidence_cell.fill = amber_fill
        else:
            confidence_cell.fill = green_fill

    for i, r in enumerate(results, 2):
        write_result_row(ws_detail, i, r)

    column_widths = {
        "A": 14, "B": 18, "C": 18, "D": 50, "E": 26,
        "F": 24, "G": 20, "H": 26, "I": 16, "J": 45,
        "K": 26, "L": 16, "M": 45, "N": 26, "O": 16,
        "P": 45, "Q": 15, "R": 40, "S": 50, "T": 50,
    }
    for col_letter, width in column_widths.items():
        ws_detail.column_dimensions[col_letter].width = width

    # --- Sheet 3: Gap Analysis ---
    ws_gap = wb.create_sheet("Gap Analysis")
    for col, h in enumerate(detail_headers, 1):
        cell = ws_gap.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    provisional_results = [
        result
        for result in results
        if is_provisional_cascade_result(result)
    ]
    gap_row = 2
    if provisional_results:
        ws_gap.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=len(detail_headers),
        )
        note_cell = ws_gap.cell(
            row=2,
            column=1,
            value=(
                f"{len(provisional_results)} provisional verdict(s) are not "
                "treated as confirmed gaps. See the Provisional Review sheet."
            ),
        )
        note_cell.font = Font(bold=True, color="8A3A22")
        note_cell.fill = amber_fill
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        gap_row = 3
    for r in results:
        if (
            r["classification"] != CLASSIFICATION_COVERS
            and not is_provisional_cascade_result(r)
        ):
            write_result_row(ws_gap, gap_row, r)
            gap_row += 1

    for col_letter, width in column_widths.items():
        ws_gap.column_dimensions[col_letter].width = width

    if provisional_results:
        ws_provisional = wb.create_sheet("Provisional Review")
        for col, header in enumerate(detail_headers, 1):
            cell = ws_provisional.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        for row_number, result in enumerate(provisional_results, 2):
            write_result_row(ws_provisional, row_number, result)
        for col_letter, width in column_widths.items():
            ws_provisional.column_dimensions[col_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_gap_analysis(results, framework_summaries):
    """Render a gap analysis summary — grouped by framework, showing only gaps."""
    provisional_count = sum(
        is_provisional_cascade_result(result) for result in results
    )
    resolved_count = len(results) - provisional_count
    gaps = [
        result
        for result in results
        if (
            result["classification"] != CLASSIFICATION_COVERS
            and not is_provisional_cascade_result(result)
        )
    ]

    if provisional_count:
        st.info(
            f"{provisional_count} provisional verdict(s) are excluded from "
            "confirmed gaps. Review them in the human review queue first."
        )

    if not gaps:
        if provisional_count and not resolved_count:
            st.warning(
                "No resolved verdicts are available yet. Review the "
                "provisional items before assessing gaps."
            )
        elif provisional_count:
            st.success(
                "No confirmed gaps were found among the resolved verdicts."
            )
        else:
            st.success(
                "No gaps found — the report covers all analysed requirements."
            )
        return

    doesnt_count = sum(1 for r in gaps if r["classification"] == CLASSIFICATION_DOESNT)
    partly_count = sum(1 for r in gaps if r["classification"] == CLASSIFICATION_PARTLY)

    st.markdown(
        f'<div style="background:#FDF6EE;border:1px solid #E6D3BE;border-radius:8px;padding:16px;margin-bottom:16px;">'
        f'<h4 style="margin:0 0 8px 0;color:#8A3A22;">Gap Analysis — What\'s Missing</h4>'
        f'<p style="margin:0;color:#3B4A40;">'
        f'<span class="badge-doesnt">{doesnt_count} not covered</span>&nbsp;&nbsp;'
        f'<span class="badge-partly">{partly_count} partly covered</span>&nbsp;&nbsp;'
        f'— these requirements need attention.</p>'
        f'</div>',
        unsafe_allow_html=True
    )

    # Group by framework
    frameworks_in_gaps = []
    for r in gaps:
        if r["framework"] not in frameworks_in_gaps:
            frameworks_in_gaps.append(r["framework"])

    for fw in frameworks_in_gaps:
        fw_gaps = [r for r in gaps if r["framework"] == fw]
        doesnt = [r for r in fw_gaps if r["classification"] == CLASSIFICATION_DOESNT]
        partly = [r for r in fw_gaps if r["classification"] == CLASSIFICATION_PARTLY]

        with st.expander(f"**{fw}** — {len(doesnt)} not covered · {len(partly)} partly covered"):
            # Not covered first (most urgent)
            if doesnt:
                st.markdown("**Not covered** — action required:")
                for r in doesnt:
                    req_text = str(r["requirement"])
                    if len(req_text) > 200:
                        req_text = req_text[:200] + "…"
                    req_text = html.escape(req_text)
                    topic_text = html.escape(
                        prettify_topic_name(str(r["topic"]))
                    )
                    ref = html.escape(str(r.get("reference", "")))
                    ref_str = f" · {ref}" if ref else ""
                    rationale = html.escape(str(r.get("rationale", "")))
                    st.markdown(
                        f'<div style="background:#F8E3DD;padding:10px;border-radius:6px;margin:6px 0;'
                        f'border-left:4px solid #B4472F;">'
                        f'<p style="margin:0 0 4px 0;font-size:13px;color:#152018;">'
                        f'<strong>[{topic_text}{ref_str}]</strong> {req_text}</p>'
                        f'<p style="margin:0;font-size:12px;color:#4B5A50;">{rationale}</p>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

            if partly:
                st.markdown("**Partly covered** — could be strengthened:")
                for r in partly:
                    req_text = str(r["requirement"])
                    if len(req_text) > 200:
                        req_text = req_text[:200] + "…"
                    req_text = html.escape(req_text)
                    topic_text = html.escape(
                        prettify_topic_name(str(r["topic"]))
                    )
                    ref = html.escape(str(r.get("reference", "")))
                    ref_str = f" · {ref}" if ref else ""
                    rationale = html.escape(str(r.get("rationale", "")))
                    st.markdown(
                        f'<div style="background:#FBF0D8;padding:10px;border-radius:6px;margin:6px 0;'
                        f'border-left:4px solid #C98A2B;">'
                        f'<p style="margin:0 0 4px 0;font-size:13px;color:#152018;">'
                        f'<strong>[{topic_text}{ref_str}]</strong> {req_text}</p>'
                        f'<p style="margin:0;font-size:12px;color:#4B5A50;">{rationale}</p>'
                        f'</div>',
                        unsafe_allow_html=True
                    )


def generate_comparison_excel(results_a, results_b, name_a, name_b, common_frameworks):
    """Generate an Excel workbook comparing two sets of results."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F3D2A")
    green_fill = PatternFill("solid", fgColor="E8F2EA")
    amber_fill = PatternFill("solid", fgColor="FBF0D8")
    red_fill = PatternFill("solid", fgColor="F8E3DD")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Framework", "Topic", "Requirement", f"{name_a}", f"{name_b}", "Difference"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Build lookup for results_b
    b_lookup = {}
    for r in results_b:
        key = (r["framework"], r["topic"], r["requirement"][:100])
        b_lookup[key] = r

    row = 2
    for r_a in results_a:
        if r_a["framework"] not in common_frameworks:
            continue
        key = (r_a["framework"], r_a["topic"], r_a["requirement"][:100])
        r_b = b_lookup.get(key)

        ws.cell(row=row, column=1, value=r_a["framework"]).border = thin_border
        ws.cell(row=row, column=2, value=prettify_topic_name(r_a["topic"])).border = thin_border
        req_text = r_a["requirement"]
        if len(req_text) > 150:
            req_text = req_text[:150] + "…"
        ws.cell(row=row, column=3, value=req_text).border = thin_border

        cell_a = ws.cell(row=row, column=4, value=r_a["classification"])
        fill_map = {CLASSIFICATION_COVERS: green_fill, CLASSIFICATION_PARTLY: amber_fill, CLASSIFICATION_DOESNT: red_fill}
        cell_a.fill = fill_map.get(r_a["classification"], PatternFill())
        cell_a.border = thin_border

        if r_b:
            cell_b = ws.cell(row=row, column=5, value=r_b["classification"])
            cell_b.fill = fill_map.get(r_b["classification"], PatternFill())
            cell_b.border = thin_border

            # Difference
            score_a = classification_to_score(r_a["classification"])
            score_b = classification_to_score(r_b["classification"])
            if score_a > score_b:
                diff = f"{name_a} better"
            elif score_b > score_a:
                diff = f"{name_b} better"
            else:
                diff = "Same"
            ws.cell(row=row, column=6, value=diff).border = thin_border
        else:
            ws.cell(row=row, column=5, value="N/A").border = thin_border
            ws.cell(row=row, column=6, value="—").border = thin_border

        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================
# REQUIREMENT-LEVEL DIFF FOR SIMILARITY VIEW
# ============================================

# Canonical topic names — the Excel contains inconsistent casings
# (e.g. 'Metricsandtargets', 'MetricsandTargets', 'Metrics and Targets')
_TOPIC_CANONICAL = {
    "metricsandtargets": "Metrics and Targets",
    "riskmanagement": "Risk Management",
    "generalrequirements": "General Requirements",
    "materialityassessment": "Materiality Assessment",
    "scenarioanalysis": "Scenario Analysis",
}


def prettify_topic_name(name):
    """Normalise topic names like 'RiskManagement' or 'Metricsandtargets' -> 'Risk Management' / 'Metrics and Targets'."""
    import re
    key = re.sub(r'[\s_\-]+', '', str(name)).lower()
    if key in _TOPIC_CANONICAL:
        return _TOPIC_CANONICAL[key]
    # Fallback: handle "and" between words, then camelCase
    spaced = re.sub(r'([a-z])(and)([A-Z])', r'\1 \2 \3', str(name))
    spaced = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', spaced)
    return spaced


def topic_match_key(name):
    """Key for matching topics across frameworks regardless of casing/spacing."""
    import re
    return re.sub(r'[\s_\-]+', '', str(name)).lower()


def compute_requirement_diffs(fw_a, fw_b, framework_requirements, requirement_refs=None):
    """
    Compare requirements between two frameworks for overlapping topics.
    Returns a list of dicts with topic, req_a, req_b, ref_a, ref_b.
    """
    import difflib

    refs = requirement_refs or {}
    reqs_a = framework_requirements.get(fw_a, {})
    reqs_b = framework_requirements.get(fw_b, {})

    # Find overlapping topic names (canonicalised: casing/spacing-insensitive)
    topics_a = {topic_match_key(t): t for t in reqs_a}
    topics_b = {topic_match_key(t): t for t in reqs_b}
    common_topics = set(topics_a.keys()) & set(topics_b.keys())

    comparisons = []
    for topic_key in sorted(common_topics):
        topic_a = topics_a[topic_key]
        topic_b = topics_b[topic_key]
        list_a = reqs_a[topic_a]
        list_b = reqs_b[topic_b]

        # Pair requirements by position (zip to shorter list)
        for i in range(max(len(list_a), len(list_b))):
            req_a_text = list_a[i] if i < len(list_a) else None
            req_b_text = list_b[i] if i < len(list_b) else None
            comparisons.append({
                "topic": topic_a,
                "req_a": req_a_text,
                "req_b": req_b_text,
                "ref_a": refs.get((fw_a, req_a_text), "") if req_a_text else "",
                "ref_b": refs.get((fw_b, req_b_text), "") if req_b_text else "",
            })

    return comparisons


def render_diff_html(text_a, text_b):
    """Generate HTML showing word-level differences between two texts."""
    import difflib

    if text_a is None:
        return (
            '<span style="color:#8A9488;font-style:italic;">'
            'No corresponding requirement</span>',
            f'<span>{text_b}</span>'
        )
    if text_b is None:
        return (
            f'<span>{text_a}</span>',
            '<span style="color:#8A9488;font-style:italic;">'
            'No corresponding requirement</span>'
        )

    words_a = text_a.split()
    words_b = text_b.split()

    # Compare on normalised tokens (lowercase, punctuation stripped) so that
    # trivial case/punctuation differences aren't highlighted — only
    # substantive word changes are.
    import re as _re

    def _norm(w):
        return _re.sub(r'[^\w]', '', w).lower()

    norm_a = [_norm(w) for w in words_a]
    norm_b = [_norm(w) for w in words_b]
    sm = difflib.SequenceMatcher(None, norm_a, norm_b)

    html_a_parts = []
    html_b_parts = []

    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == 'equal':
            html_a_parts.append(" ".join(words_a[i1:i2]))
            html_b_parts.append(" ".join(words_b[j1:j2]))
        elif op == 'replace':
            html_a_parts.append(
                f'<span style="background:#EEC2B4;padding:1px 3px;border-radius:3px;">'
                f'{" ".join(words_a[i1:i2])}</span>'
            )
            html_b_parts.append(
                f'<span style="background:#EEC2B4;padding:1px 3px;border-radius:3px;">'
                f'{" ".join(words_b[j1:j2])}</span>'
            )
        elif op == 'delete':
            html_a_parts.append(
                f'<span style="background:#EEC2B4;padding:1px 3px;border-radius:3px;">'
                f'{" ".join(words_a[i1:i2])}</span>'
            )
        elif op == 'insert':
            html_b_parts.append(
                f'<span style="background:#EEC2B4;padding:1px 3px;border-radius:3px;">'
                f'{" ".join(words_b[j1:j2])}</span>'
            )

    return " ".join(html_a_parts), " ".join(html_b_parts)


# ============================================
# MAIN APP
# ============================================


def main():
    st.title("Sustainability Framework Analyser")
    st.markdown("Compare & analyse ESG reporting frameworks")

    # Load requirements + source references from Excel once
    framework_requirements, requirement_refs = load_framework_requirements()

    # Load similarity CSVs
    similarity_data = load_similarity_data()

    tab0, tab1, tab2, tab3, tab4 = st.tabs([
        "Welcome", "Framework Map", "Report Analyser",
        "Side-by-Side Comparison", "Report Drafter (Beta)"
    ])

    # ============================================
    # TAB 0: WELCOME / INTRODUCTION
    # ============================================
    with tab0:
        # ── Hero ──
        n_frameworks = len(FRAMEWORK_COLORS)
        n_countries = len({c for ctrs in ADOPTION_DICT.values() for c in ctrs})
        n_requirements = sum(
            len(reqs)
            for topics in framework_requirements.values()
            for reqs in topics.values()
        )
        chips_html = "".join(
            f'<span style="display:inline-flex;align-items:center;gap:8px;'
            f'background:#FCFAF3;border:1px solid #DDD5C2;border-radius:9px;'
            f'padding:8px 14px;font-weight:600;font-size:14px;color:#152018;'
            f'margin:0 8px 8px 0;" title="{FRAMEWORK_FULL_NAMES.get(fw, fw)}">'
            f'<span style="width:9px;height:9px;border-radius:50%;'
            f'background:{color};display:inline-block;flex-shrink:0;"></span>'
            f'{fw}</span>'
            for fw, color in FRAMEWORK_COLORS.items()
        )
        hero_left, hero_right = st.columns([1.05, 0.95], gap="large")
        with hero_left:
            st.markdown(
                f'''
<div style="padding:26px 0 6px;">
  <div style="display:inline-flex;align-items:center;gap:8px;background:#E8F2EA;border:1px solid #C6E0CC;border-radius:20px;padding:6px 14px;font-size:13px;font-weight:600;color:#1C6B4A;margin-bottom:22px;"><span style="width:7px;height:7px;border-radius:50%;background:#1C6B4A;display:inline-block;"></span> AI-assisted &middot; requirement-level analysis</div>
  <h1 style="font-family:'Spectral',serif;font-weight:600;font-size:48px;line-height:1.06;letter-spacing:-0.01em;margin:0 0 18px;color:#152018;">Know exactly where your ESG report stands.</h1>
  <p style="font-size:17px;line-height:1.6;color:#4B5A50;margin:0 0 28px;max-width:560px;">Upload a transition plan or sustainability report and see how it measures against <strong style="color:#152018;">{n_frameworks} global disclosure frameworks</strong> — assessed requirement by requirement, with the exact passages that support each finding.</p>
  <div style="display:flex;gap:44px;">
    <div><div style="font-family:'Spectral',serif;font-size:36px;font-weight:600;color:#0F3D2A;line-height:1;">{n_frameworks}</div><div style="font-size:13px;color:#8A9488;font-weight:500;margin-top:5px;">frameworks tracked</div></div>
    <div><div style="font-family:'Spectral',serif;font-size:36px;font-weight:600;color:#0F3D2A;line-height:1;">{n_countries}</div><div style="font-size:13px;color:#8A9488;font-weight:500;margin-top:5px;">countries mapped</div></div>
    <div><div style="font-family:'Spectral',serif;font-size:36px;font-weight:600;color:#0F3D2A;line-height:1;">{n_requirements}</div><div style="font-size:13px;color:#8A9488;font-weight:500;margin-top:5px;">requirements assessed</div></div>
  </div>
</div>
''',
                unsafe_allow_html=True,
            )
        with hero_right:
            st.markdown(
                '''
<div style="background:#FCFAF3;border:1px solid #DDD5C2;border-radius:16px;padding:24px;margin-top:26px;box-shadow:0 18px 40px -24px rgba(30,40,30,.35);">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;"><span style="font-weight:700;font-size:14px;color:#152018;">Coverage summary</span><span style="font-family:'IBM Plex Mono',monospace;font-size:10.5px;color:#8A9488;">example-report.pdf</span></div>
  <div style="display:flex;align-items:center;gap:20px;margin-bottom:20px;">
    <div style="width:96px;height:96px;border-radius:50%;background:conic-gradient(#1C6B4A 0 68%, #DDD5C2 68% 100%);display:flex;align-items:center;justify-content:center;flex-shrink:0;"><div style="width:70px;height:70px;border-radius:50%;background:#FCFAF3;display:flex;flex-direction:column;align-items:center;justify-content:center;"><span style="font-family:'Spectral',serif;font-size:24px;font-weight:600;color:#1C6B4A;line-height:1;">68%</span><span style="font-size:9px;color:#8A9488;font-weight:500;">overall</span></div></div>
    <div style="flex:1;">
      <div style="display:flex;align-items:center;gap:9px;margin-bottom:8px;"><span style="width:10px;height:10px;border-radius:3px;background:#1C6B4A;display:inline-block;"></span><span style="font-size:12.5px;color:#4B5A50;flex:1;">Covers</span><span style="font-weight:700;font-size:12.5px;color:#152018;">142</span></div>
      <div style="display:flex;align-items:center;gap:9px;margin-bottom:8px;"><span style="width:10px;height:10px;border-radius:3px;background:#C98A2B;display:inline-block;"></span><span style="font-size:12.5px;color:#4B5A50;flex:1;">Partly</span><span style="font-weight:700;font-size:12.5px;color:#152018;">68</span></div>
      <div style="display:flex;align-items:center;gap:9px;"><span style="width:10px;height:10px;border-radius:3px;background:#B4472F;display:inline-block;"></span><span style="font-size:12.5px;color:#4B5A50;flex:1;">Doesn't</span><span style="font-weight:700;font-size:12.5px;color:#152018;">31</span></div>
    </div>
  </div>
  <div style="margin-bottom:12px;"><div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px;"><span style="font-weight:600;color:#152018;">TCFD</span><span style="font-family:'IBM Plex Mono',monospace;color:#8A9488;">81%</span></div><div style="height:8px;border-radius:5px;background:#DDD5C2;overflow:hidden;display:flex;"><div style="width:81%;background:#1C6B4A;"></div></div></div>
  <div style="margin-bottom:12px;"><div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px;"><span style="font-weight:600;color:#152018;">IFRS S2</span><span style="font-family:'IBM Plex Mono',monospace;color:#8A9488;">72%</span></div><div style="height:8px;border-radius:5px;background:#DDD5C2;overflow:hidden;display:flex;"><div style="width:72%;background:#1C6B4A;"></div></div></div>
  <div style="margin-bottom:12px;"><div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px;"><span style="font-weight:600;color:#152018;">TNFD</span><span style="font-family:'IBM Plex Mono',monospace;color:#8A9488;">54%</span></div><div style="height:8px;border-radius:5px;background:#DDD5C2;overflow:hidden;display:flex;"><div style="width:54%;background:#C98A2B;"></div></div></div>
  <div><div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px;"><span style="font-weight:600;color:#152018;">ESRS E1</span><span style="font-family:'IBM Plex Mono',monospace;color:#8A9488;">63%</span></div><div style="height:8px;border-radius:5px;background:#DDD5C2;overflow:hidden;display:flex;"><div style="width:63%;background:#1C6B4A;"></div></div></div>
  <p style="margin:14px 0 0;font-family:'IBM Plex Mono',monospace;font-size:10px;color:#8A9488;text-align:right;">example output</p>
</div>
''',
                unsafe_allow_html=True,
            )
        st.markdown(
            f'<div style="margin:10px 0 14px;">{chips_html}</div>'
            f'<p style="font-size:13px;color:#8A9488;margin:0 0 6px;">Built by the '
            f'<strong style="color:#4B5A50;">IFoA Sustainability and Reporting '
            f'Working Party</strong>.</p>',
            unsafe_allow_html=True,
        )

        st.markdown("### What you can do")
        feature_cols = st.columns(3)
        feature_cards = [
            (
                "Framework Map",
                "Explore global adoption on the interactive globe and compare "
                "similarity between frameworks across governance, strategy, "
                "risk, metrics and disclosure.",
            ),
            (
                "Report Analyser",
                "Upload a PDF and have it assessed requirement-by-requirement. "
                "Each one is classified as Covered, Partly covered or Not "
                "covered, with rationale and extracts.",
            ),
            (
                "Side-by-Side Comparison",
                "Benchmark two reports against each other — useful for "
                "year-on-year progress or comparing two firms' disclosures.",
            ),
        ]
        for col, (card_title, card_body) in zip(feature_cols, feature_cards):
            with col:
                st.markdown(
                    f'<div style="background:#FCFAF3;border:1px solid #DDD5C2;'
                    f'border-radius:12px;padding:20px;min-height:150px;">'
                    f'<p style="margin:0 0 8px;font-size:16px;font-weight:700;'
                    f'color:#152018;">{card_title}</p>'
                    f'<p style="margin:0;font-size:13.5px;line-height:1.55;'
                    f'color:#4B5A50;">{card_body}</p>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("### How to use the Report Analyser")
        st.markdown(
            "1. Go to the **Report Analyser** tab.\n"
            "2. Select the frameworks you want to assess against.\n"
            "3. Upload your PDF and optionally select a page range.\n"
            "4. Click **Analyse Report** and wait for results.\n"
            "5. Review the detailed analysis, gap analysis, and download the Excel export."
        )

        st.markdown("---")
        st.markdown("### The Team")

        # ── Team profiles ──
        # Edit the list below to add/remove members.
        team_members = [
            {
                "name": "Lloyd Richards",
                "title": "Director & Head of Actuarial",
                "organisation": "Crowe UK",
                "linkedin_url": "https://www.linkedin.com/in/lloydrichards/",
                "bio": "Lead of the IFoA Sustainability and Reporting Working Party.",
            },
            {
                "name": "Cristian-Anton Calin",
                "title": "Pricing Actuary",
                "organisation": "Zurich Insurance",
                "linkedin_url": "https://linkedin.com/in/cristian-calin-b4969b1b2",
                "bio": "App Technical Developer.",
            },
            {
                "name": "Charchit Agrawal",
                "title": "Associate Director",
                "organisation": "BDO",
                "linkedin_url": "https://www.linkedin.com/in/charchit-agrawal-ba2a334/",
                "bio": "...",
            },
            {
                "name": "Diksha Thawani",
                "title": "Student Actuary",
                "organisation": "Zurich Kotak General Insurance",
                "linkedin_url": "https://www.linkedin.com/in/diksha-thawani/",
                "bio": "...",
            },
            {
                "name": "Ella Reid-Norris",
                "title": "Actuary, economist and strategist",
                "organisation": "Fair4All Finance",
                "linkedin_url": "https://www.linkedin.com/in/ella-reid-norris/",
                "bio": "...",
            },
            {
                "name": "Festus Cheruiyot",
                "title": "Actuarial Student",
                "organisation": "ACTEX Learning",
                "linkedin_url": "https://www.linkedin.com/in/festus-cheruiyot/",
                "bio": "...",
            },
            {
                "name": "Stephen Goh",
                "title": "Actuarial Consultant",
                "organisation": "Milliman UK",
                "linkedin_url": "https://www.linkedin.com/in/stephengoh/",
                "bio": "...",
            },
        ]

        cols_per_row = 3
        for row_start in range(0, len(team_members), cols_per_row):
            row_members = team_members[row_start:row_start + cols_per_row]
            profile_cols = st.columns(cols_per_row)
            for idx, member in enumerate(row_members):
                with profile_cols[idx]:
                    org_line = ""
                    if member["organisation"]:
                        org_line = (
                            '<p style="margin:0 0 8px 0;font-size:12px;'
                            f'color:#8A9488;">{member["organisation"]}</p>'
                        )
                    st.markdown(
                        f'<div style="background:#EDE7D8;border:1px solid #DDD5C2;'
                        f'border-radius:10px;padding:20px;min-height:180px;">'
                        f'<p style="margin:0;font-size:17px;font-weight:700;'
                        f'color:#152018;">{member["name"]}</p>'
                        f'<p style="margin:2px 0 4px 0;font-size:13px;'
                        f'color:#C98A2B;font-weight:600;">{member["title"]}</p>'
                        f'{org_line}'
                        f'<p style="margin:0 0 10px 0;font-size:13px;'
                        f'color:#4B5A50;">{member["bio"]}</p>'
                        f'<a href="{member["linkedin_url"]}" target="_blank" '
                        f'style="font-size:13px;color:#1C6B4A;text-decoration:none;">'
                        f'LinkedIn &rarr;</a>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

    # ============================================
    # TAB 1: FRAMEWORK MAP
    # ============================================
    with tab1:
        st.header("Climate & Sustainability Framework Adoption")
        st.markdown(
            "Explore global adoption and compare similarity between "
            "regulatory frameworks"
        )

        # --- Controls ---
        ctrl_col1, ctrl_col2, ctrl_col3 = st.columns([1, 1, 2])

        with ctrl_col1:
            framework_options = ["ALL"] + list(FRAMEWORK_COLORS.keys())
            selected_framework = st.selectbox(
                "Select Framework",
                options=framework_options,
                help=(
                    "Choose a framework to see its similarity to others, "
                    "or ALL for the full map"
                )
            )

        with ctrl_col2:
            available_metrics = (
                list(similarity_data.keys())
                if similarity_data
                else SIMILARITY_METRIC_TYPES
            )
            metric_type = st.selectbox(
                "Select Metric Type",
                options=available_metrics,
                format_func=lambda x: x.replace("_", " ").title(),
                help="Filter similarity scores by topic area"
            )

        # --- Main content: Similarity table (prominent) + Legend ---
        content_col, legend_col = st.columns([3, 1])

        with legend_col:
            st.markdown("#### Framework Legend")
            st.caption(
                "(n) = number of adopting jurisdictions. "
                "Click a name to visit the official source. "
                "Hover for the full framework name."
            )
            for fw, color in FRAMEWORK_COLORS.items():
                full_name = FRAMEWORK_FULL_NAMES.get(fw, fw)
                count = len(ADOPTION_DICT.get(fw, []))
                fw_url = FRAMEWORK_URLS.get(fw, "")
                name_html = (
                    f'<a href="{fw_url}" target="_blank" '
                    f'style="color:#152018;font-size:13px;'
                    f'text-decoration:underline;text-decoration-color:#CFC7B2;">{fw}</a>'
                    if fw_url else
                    f'<span style="color:#152018;font-size:13px;">{fw}</span>'
                )
                st.markdown(
                    f'<div style="display:flex;align-items:center;gap:8px;'
                    f'margin:4px 0;" title="{full_name}">'
                    f'<div style="width:14px;height:14px;background:{color};'
                    f'border-radius:3px;flex-shrink:0;"></div>'
                    f'{name_html}'
                    f'<span style="color:#8A9488;font-size:12px;">({count})</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )

        with content_col:
            if selected_framework == "ALL":
                st.markdown(
                    '<div style="background:#EDE7D8;border:1px solid #DDD5C2;'
                    'border-radius:8px;padding:40px;text-align:center;'
                    'margin:16px 0;">'
                    '<p style="color:#4B5A50;font-size:15px;margin:0;">'
                    '&#x1F446; Select a framework from the dropdown above to '
                    'explore its similarity to other frameworks.</p>'
                    '</div>',
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f"#### Framework Similarity: {selected_framework}"
                )
                st.markdown(
                    f"*{FRAMEWORK_FULL_NAMES.get(selected_framework, selected_framework)}"
                    f" · Metric: {metric_type.replace('_', ' ').title()}*"
                )
                with st.expander("ℹ️ How are these scores calculated?"):
                    st.markdown(
                        "**Similarity scores** are computed by converting each framework's "
                        "requirement texts into sentence embeddings (numerical vectors capturing "
                        "meaning) and measuring the cosine similarity between frameworks. "
                        "A higher percentage means the two frameworks ask for more semantically "
                        "similar things.\n\n"
                        "**Metric Type** filters which *dimension* of the pre-computed similarity "
                        "is shown (e.g. 'Risk' compares only risk-related content). The topic "
                        "headings you see inside the requirement-level comparison (Governance, "
                        "Strategy, Risk Management, Metrics and Targets, ...) come from each "
                        "framework's *own document structure*, so they are related to but not "
                        "identical to the Metric Type filter."
                    )

                df_sim = similarity_data.get(metric_type)
                if df_sim is None:
                    st.info(
                        f"No similarity data available for "
                        f"{metric_type.replace('_', ' ').title()}. "
                        f"Check that the CSV file is present in the repository."
                    )
                    similarities = []
                else:
                    similarities = get_similarity_for_framework(
                        df_sim, selected_framework
                    )

                if similarities:
                    for item in similarities:
                        score = item['similarity']
                        pct = score * 100
                        other_fw = item['framework']
                        color = (
                            "#1C6B4A" if score >= 0.4
                            else "#4E8A67" if score >= 0.3
                            else "#C98A2B" if score >= 0.2
                            else "#B4472F"
                        )
                        fw_color = FRAMEWORK_COLORS.get(other_fw, "#8A9488")
                        other_full = FRAMEWORK_FULL_NAMES.get(other_fw, other_fw)

                        st.markdown(
                            f'<div style="background:#EDE7D8;padding:12px;'
                            f'border-radius:8px;margin:8px 0;">'
                            f'<div style="display:flex;justify-content:space-between;'
                            f'align-items:center;">'
                            f'<div style="display:flex;align-items:center;gap:8px;">'
                            f'<div style="width:14px;height:14px;background:{fw_color};'
                            f'border-radius:3px;"></div>'
                            f'<span style="font-weight:600;color:#152018;" '
                            f'title="{other_full}">{other_fw}</span>'
                            f'</div>'
                            f'<span style="color:{color};font-weight:700;'
                            f'font-family:monospace;">{pct:.1f}%</span>'
                            f'</div>'
                            f'<div style="background:#DDD5C2;border-radius:4px;'
                            f'height:8px;margin-top:8px;overflow:hidden;">'
                            f'<div style="background:{color};height:100%;'
                            f'width:{pct}%;"></div>'
                            f'</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                        # Expandable requirement-level comparison.
                        # Similarity data may use parent names (ESRS, IFRS) while
                        # requirements use split names (ESRS E1/E4, IFRS S1/S2) —
                        # resolve to every matching split.
                        if other_fw in framework_requirements:
                            diff_targets = [other_fw]
                        else:
                            diff_targets = [
                                split for split, parent in SIMILARITY_PARENT_MAP.items()
                                if parent == other_fw
                                and split in framework_requirements
                            ]

                        for diff_fw in diff_targets:
                            comparisons = compute_requirement_diffs(
                                selected_framework, diff_fw,
                                framework_requirements, requirement_refs
                            )
                            if comparisons:
                                with st.expander(
                                    f"View requirement-level comparison: "
                                    f"{selected_framework} vs {diff_fw} "
                                    f"({len(comparisons)} requirement pairs)"
                                ):
                                    current_topic = None
                                    for comp in comparisons:
                                        if comp["topic"] != current_topic:
                                            current_topic = comp["topic"]
                                            st.markdown(f"**{prettify_topic_name(current_topic)}**")

                                        html_a, html_b = render_diff_html(
                                            comp["req_a"], comp["req_b"]
                                        )
                                        ref_a_html = (
                                            f' <span style="font-size:10px;color:#8A9488;'
                                            f'font-family:monospace;">({comp["ref_a"]})</span>'
                                            if comp.get("ref_a") else ""
                                        )
                                        ref_b_html = (
                                            f' <span style="font-size:10px;color:#8A9488;'
                                            f'font-family:monospace;">({comp["ref_b"]})</span>'
                                            if comp.get("ref_b") else ""
                                        )

                                        fw_sel_color = FRAMEWORK_COLORS.get(
                                            selected_framework, "#8A9488"
                                        )
                                        diff_fw_color = FRAMEWORK_COLORS.get(
                                            diff_fw, fw_color
                                        )
                                        st.markdown(
                                            f'<div style="display:flex;gap:12px;'
                                            f'margin:6px 0;font-size:12px;'
                                            f'line-height:1.5;">'
                                            f'<div style="flex:1;background:#F1EDE1;'
                                            f'padding:8px;border-radius:6px;'
                                            f'border-left:3px solid {fw_sel_color};">'
                                            f'<strong style="color:#152018;">'
                                            f'{selected_framework}</strong>{ref_a_html}'
                                            f'<br><span style="color:#3B4A40;">'
                                            f'{html_a}</span></div>'
                                            f'<div style="flex:1;background:#EDF3EA;'
                                            f'padding:8px;border-radius:6px;'
                                            f'border-left:3px solid {diff_fw_color};">'
                                            f'<strong style="color:#152018;">'
                                            f'{diff_fw}</strong>{ref_b_html}'
                                            f'<br><span style="color:#3B4A40;">'
                                            f'{html_b}</span></div>'
                                            f'</div>',
                                            unsafe_allow_html=True
                                        )
                else:
                    st.info(
                        f"No similarity data available for "
                        f"{selected_framework} under "
                        f"{metric_type.replace('_', ' ').title()}"
                    )

        # --- Globe map below ---
        st.markdown("---")
        st.markdown("#### Global Adoption Map")

        map_data = []
        if selected_framework == "ALL":
            all_countries = set()
            for countries_list in ADOPTION_DICT.values():
                all_countries.update(countries_list)
            for country in all_countries:
                if country in COUNTRY_COORDS:
                    fws = [
                        fw for fw, ctrs in ADOPTION_DICT.items()
                        if country in ctrs
                    ]
                    map_data.append({
                        "country": country,
                        "lat": COUNTRY_COORDS[country]["lat"],
                        "lon": COUNTRY_COORDS[country]["lon"],
                        "frameworks": len(fws),
                        "framework_list": ", ".join(fws),
                        "size": 10 + len(fws) * 3
                    })
        else:
            for country in ADOPTION_DICT.get(selected_framework, []):
                if country in COUNTRY_COORDS:
                    map_data.append({
                        "country": country,
                        "lat": COUNTRY_COORDS[country]["lat"],
                        "lon": COUNTRY_COORDS[country]["lon"],
                        "frameworks": 1,
                        "framework_list": selected_framework,
                        "size": 15
                    })

        if map_data:
            df_map = pd.DataFrame(map_data)
            fig = px.scatter_geo(
                df_map, lat="lat", lon="lon",
                hover_name="country",
                hover_data={
                    "framework_list": True, "lat": False,
                    "lon": False, "frameworks": False, "size": False
                },
                size="size",
                color="frameworks" if selected_framework == "ALL" else None,
                color_continuous_scale=(
                    ["#E3DCC5", "#4E8A67", "#0F3D2A"] if selected_framework == "ALL" else None
                ),
                projection="orthographic"
            )

            if selected_framework != "ALL":
                fig.update_traces(marker=dict(
                    color=FRAMEWORK_COLORS.get(selected_framework, "#8A9488")
                ))

            n_frames = 36
            frames = []
            for i in range(n_frames):
                lon_rot = -20 + (360 / n_frames) * i
                frames.append(go.Frame(
                    layout=dict(
                        geo=dict(projection_rotation=dict(lon=lon_rot, lat=15))
                    ),
                    name=str(i)
                ))
            fig.frames = frames

            fig.update_layout(
                geo=dict(
                    showland=True, landcolor="#DCE3CC",
                    showocean=True, oceancolor="#EDE7D8",
                    showcoastlines=True, coastlinecolor="#B4AE9C",
                    showcountries=True, countrycolor="#CBC3AE",
                    showframe=False, bgcolor="#F5F1E8",
                    projection_rotation=dict(lon=-20, lat=15),
                ),
                paper_bgcolor="#F5F1E8", plot_bgcolor="#F5F1E8",
                font=dict(color="#152018", family="Hanken Grotesk, sans-serif"),
                margin=dict(l=0, r=0, t=0, b=0), height=500,
                updatemenus=[dict(
                    type="buttons", showactive=False,
                    x=0.02, y=0.02, xanchor="left", yanchor="bottom",
                    buttons=[
                        dict(
                            label="\u25b6 Spin", method="animate",
                            args=[None, dict(
                                frame=dict(duration=120, redraw=True),
                                fromcurrent=True,
                                transition=dict(duration=80),
                                mode="immediate",
                            )],
                        ),
                        dict(
                            label="\u23f8 Stop", method="animate",
                            args=[[None], dict(
                                frame=dict(duration=0, redraw=False),
                                mode="immediate",
                                transition=dict(duration=0),
                            )],
                        ),
                    ],
                )],
            )
            st.plotly_chart(fig, use_container_width=True)

    # ============================================
    # TAB 2: REPORT ANALYSER (single-column)
    # ============================================
    with tab2:
        pending_analysis = st.session_state.get("pending_analysis")
        pending_batch_id = (
            pending_analysis.get("batch_id") if pending_analysis else None
        )
        pending_model_id = (
            pending_analysis.get("model_id", PRIMARY_MODEL)
            if pending_analysis
            else PRIMARY_MODEL
        )
        if pending_model_id not in USER_SELECTABLE_MODELS:
            pending_model_id = PRIMARY_MODEL
        if pending_batch_id:
            st.session_state["analysis_model_id"] = pending_model_id
            st.session_state["analysis_strategy"] = ANALYSIS_STRATEGY_SINGLE

        st.header("ESG Report Analyser")
        st.markdown(
            "Upload your transition plan or ESG report PDF to analyse how "
            "well it aligns with sustainability frameworks. Use one chosen "
            "model, or select the reviewed cascade for explicit model review "
            "and conditional adjudication."
        )

        # --- All controls on one row ---
        st.markdown("---")
        api_col, upload_col = st.columns([1, 1])

        with api_col:
            analysis_strategy = st.radio(
                "Analysis strategy",
                options=[
                    ANALYSIS_STRATEGY_SINGLE,
                    ANALYSIS_STRATEGY_CASCADE,
                ],
                index=0,
                key="analysis_strategy",
                disabled=bool(pending_batch_id),
                horizontal=True,
                help=(
                    "Choose an analyst and reviewer for every requirement, "
                    "plus a senior reviewer that is called only when their "
                    "classifications disagree."
                ),
            )
            is_review_cascade = (
                analysis_strategy == ANALYSIS_STRATEGY_CASCADE
            )
            anthropic_api_key = ""
            openai_api_key = ""
            api_key = ""
            required_cascade_providers = set()

            if is_review_cascade:
                analyst_model_id = st.selectbox(
                    "Analyst",
                    options=ANALYST_MODELS,
                    index=ANALYST_MODELS.index(HAIKU_MODEL),
                    format_func=model_picker_label,
                    key="cascade_analyst_model_id",
                    help="The analyst performs the initial assessment.",
                )
                available_reviewers = tuple(
                    model_id
                    for model_id in REVIEWER_MODELS
                    if model_id != analyst_model_id
                )
                if (
                    st.session_state.get("cascade_reviewer_model_id")
                    not in available_reviewers
                ):
                    st.session_state["cascade_reviewer_model_id"] = (
                        LUNA_MODEL
                        if LUNA_MODEL in available_reviewers
                        else available_reviewers[0]
                    )
                reviewer_model_id = st.selectbox(
                    "Reviewer",
                    options=available_reviewers,
                    format_func=model_picker_label,
                    key="cascade_reviewer_model_id",
                    help=(
                        "The reviewer independently checks every analyst "
                        "assessment. The analyst model is excluded."
                    ),
                )
                available_senior_reviewers = tuple(
                    model_id
                    for model_id in SENIOR_REVIEWER_MODELS
                    if model_id != reviewer_model_id
                )
                if (
                    st.session_state.get("cascade_senior_reviewer_model_id")
                    not in available_senior_reviewers
                ):
                    st.session_state["cascade_senior_reviewer_model_id"] = (
                        TERRA_MODEL
                        if TERRA_MODEL in available_senior_reviewers
                        else available_senior_reviewers[0]
                    )
                senior_reviewer_model_id = st.selectbox(
                    "Senior reviewer (disagreements only)",
                    options=available_senior_reviewers,
                    format_func=model_picker_label,
                    key="cascade_senior_reviewer_model_id",
                    help=(
                        "Called only when analyst and reviewer classifications "
                        "differ. The reviewer model is excluded."
                    ),
                )
                selected_model_id = analyst_model_id
                selected_cascade_models = (
                    analyst_model_id,
                    reviewer_model_id,
                    senior_reviewer_model_id,
                )
                required_cascade_providers = {
                    get_model_config(model_id)["provider"]
                    for model_id in selected_cascade_models
                }
                analyst_label = get_model_config(analyst_model_id)["label"]
                reviewer_label = get_model_config(reviewer_model_id)["label"]
                senior_label = get_model_config(
                    senior_reviewer_model_id
                )["label"]
                st.warning(
                    "**Reviewed cascade is slower and costlier.** "
                    f"{analyst_label} and {reviewer_label} run for every "
                    f"requirement. {senior_label} is charged only when their "
                    "classifications disagree. The same model cannot occupy "
                    "adjacent roles. Sequential review uses standard API "
                    "pricing."
                )
                st.markdown("**Required credentials**")
                if "anthropic" in required_cascade_providers:
                    anthropic_model = next(
                        model_id
                        for model_id in selected_cascade_models
                        if get_model_config(model_id)["provider"] == "anthropic"
                    )
                    anthropic_api_key = render_model_api_key(
                        anthropic_model, "analysis_anthropic_api_key"
                    )
                if "openai" in required_cascade_providers:
                    openai_model = next(
                        model_id
                        for model_id in selected_cascade_models
                        if get_model_config(model_id)["provider"] == "openai"
                    )
                    openai_api_key = render_model_api_key(
                        openai_model, "analysis_openai_api_key"
                    )
                with st.expander("Model costs in this cascade"):
                    st.markdown(f"**Analyst — {analyst_label}**")
                    render_model_price_caption(analyst_model_id)
                    st.markdown(f"**Reviewer — {reviewer_label}**")
                    render_model_price_caption(reviewer_model_id)
                    st.markdown(
                        f"**Senior reviewer — {senior_label} "
                        "(disagreements only)**"
                    )
                    render_model_price_caption(senior_reviewer_model_id)
            else:
                selected_model_id = st.selectbox(
                    "Analysis model",
                    options=USER_SELECTABLE_MODELS,
                    index=0,
                    format_func=model_picker_label,
                    key="analysis_model_id",
                    disabled=bool(pending_batch_id),
                    help=(
                        "A pending batch locks its original provider and model "
                        "until it is resumed or cleared."
                    ),
                )
                render_model_price_caption(selected_model_id)
                selected_provider = get_model_config(
                    selected_model_id
                )["provider"]
                api_key = render_model_api_key(
                    selected_model_id,
                    f"analysis_{selected_provider}_api_key",
                )

            st.markdown("**Select Frameworks**")
            available_frameworks = (
                list(framework_requirements.keys())
                if framework_requirements
                else list(FRAMEWORK_COLORS.keys())
            )

            btn_col1, btn_col2 = st.columns(2)
            with btn_col1:
                if st.button("Select All"):
                    st.session_state.selected_frameworks = (
                        available_frameworks.copy()
                    )
                    for framework in available_frameworks:
                        st.session_state[f"fw_{framework}"] = True
            with btn_col2:
                if st.button("Clear All"):
                    st.session_state.selected_frameworks = []
                    for framework in available_frameworks:
                        st.session_state[f"fw_{framework}"] = False

            if 'selected_frameworks' not in st.session_state:
                st.session_state.selected_frameworks = ["TCFD", "TNFD"]

            selected_frameworks = []
            fw_cols = st.columns(3)
            for i, fw in enumerate(available_frameworks):
                with fw_cols[i % 3]:
                    req_count = sum(
                        len(reqs)
                        for reqs in framework_requirements.get(fw, {}).values()
                    )
                    checked = st.checkbox(
                        f"{fw}",
                        value=fw in st.session_state.selected_frameworks,
                        key=f"fw_{fw}",
                        help=(
                            f"{FRAMEWORK_FULL_NAMES.get(fw, fw)} "
                            f"({req_count} requirements)"
                        )
                    )
                    if checked:
                        selected_frameworks.append(fw)

            st.session_state.selected_frameworks = selected_frameworks

            total_reqs = sum(
                sum(
                    len(reqs)
                    for reqs in framework_requirements.get(fw, {}).values()
                )
                for fw in selected_frameworks
            )
            st.markdown(
                f"**{len(selected_frameworks)}** framework(s) selected "
                f"\u00b7 **{total_reqs}** requirements"
            )

        with upload_col:
            st.markdown("**Upload Document**")
            if is_review_cascade:
                st.info(
                    "**Processing time:** Reviewed cascade makes two full "
                    "sequential passes, plus conditional senior-reviewer "
                    "calls. Higher-capability role choices can take "
                    "substantially longer. Vision also adds time because page "
                    "images must be rendered, uploaded, and analysed."
                )
            else:
                st.info(
                    "**Processing time:** Batch API is the cheaper but slower "
                    "option and may take minutes or substantially longer. Turn "
                    "it off for the fastest interactive result. Vision also "
                    "adds processing time because page images must be rendered, "
                    "uploaded, and analysed."
                )
            use_vision = st.checkbox(
                "Use vision for charts and image-based tables — slower",
                value=True,
                help=(
                    "Renders up to 30 visually dense or scanned pages and sends "
                    "them to the selected model alongside page-tagged text. "
                    "This improves coverage of visual disclosures but increases "
                    "upload and analysis time."
                ),
            )
            if is_review_cascade:
                use_batch_api = st.checkbox(
                    "Batch API unavailable for reviewed cascade",
                    value=False,
                    disabled=True,
                    key="analysis_cascade_batch_disabled",
                    help=(
                        "The review stages depend on earlier verdicts, so this "
                        "mode uses sequential standard API calls."
                    ),
                )
                st.caption(
                    "Sequential review uses standard API calls; Batch API "
                    "remains available in Single model mode."
                )
            else:
                use_batch_api = st.checkbox(
                    "Use Batch API — 50% cheaper, but slower",
                    value=True,
                    key="analysis_single_use_batch",
                    help=(
                        "Uses Anthropic Message Batches or OpenAI Batch, "
                        "depending on the selected model. Batch is designed for "
                        "lower cost, not immediate results: it may take minutes "
                        "and can take up to 24 hours. Turn it off for faster "
                        "interactive testing."
                    ),
                )
            uploaded_file = st.file_uploader(
                "Choose a PDF file", type="pdf",
                help="Upload your ESG report or transition plan PDF"
            )

            page_start = 1
            page_end = None
            if uploaded_file:
                import pymupdf
                pdf_bytes = uploaded_file.read()
                uploaded_file.seek(0)
                with pymupdf.open(stream=pdf_bytes, filetype="pdf") as doc:
                    total_pages = len(doc)

                st.markdown(
                    f"**PDF has {total_pages} pages.** "
                    f"Select the range to analyse:"
                )
                pr_col1, pr_col2 = st.columns(2)
                with pr_col1:
                    page_start = st.number_input(
                        "From page", min_value=1, max_value=total_pages,
                        value=1, step=1, key="page_start"
                    )
                with pr_col2:
                    page_end = st.number_input(
                        "To page", min_value=1, max_value=total_pages,
                        value=total_pages, step=1, key="page_end"
                    )
                if page_start > page_end:
                    st.warning("'From page' must be \u2264 'To page'.")

            st.markdown("**Or paste text:**")
            pasted_text = st.text_area(
                "Paste your report text here", height=120,
                placeholder="Paste your ESG report content..."
            )

        # A submitted batch can outlive a Streamlit run. Keep enough context in
        # session state to retrieve it without creating and paying for another.
        pending_analysis = st.session_state.get("pending_analysis")
        pending_batch_id = (
            pending_analysis.get("batch_id") if pending_analysis else None
        )
        resume_pending = False
        if pending_batch_id:
            pending_label = get_model_config(pending_model_id)["label"]
            st.warning(
                f"Batch `{pending_batch_id}` for {pending_label} is still "
                "available. Resume it instead of submitting the report again."
            )
            resume_col, clear_col = st.columns(2)
            with resume_col:
                resume_pending = st.button(
                    "Resume Pending Batch", type="primary", disabled=not api_key
                )
            with clear_col:
                if st.button("Clear Pending Reference"):
                    st.session_state.pop("pending_analysis", None)
                    pending_analysis = None
                    pending_batch_id = None

        if resume_pending and pending_analysis:
            st.markdown("### Retrieving Batch...")
            resume_progress = st.progress(0)
            try:
                results, framework_summaries, token_usage = (
                    run_model_analysis(
                        pending_analysis["report_text"],
                        pending_analysis["selected_frameworks"],
                        api_key,
                        framework_requirements,
                        resume_progress,
                        requirement_refs,
                        report_pages=pending_analysis["report_pages"],
                        use_batch=True,
                        existing_batch_id=pending_batch_id,
                        track_pending_batch=True,
                        model_id=pending_model_id,
                    )
                )
                st.session_state.analysis_results = results
                st.session_state.framework_summaries = framework_summaries
                st.session_state.num_pages = pending_analysis["num_pages"]
                st.session_state.token_usage = token_usage
                st.session_state.analysis_strategy_used = (
                    ANALYSIS_STRATEGY_SINGLE
                )
                st.session_state.selected_frameworks = pending_analysis[
                    "selected_frameworks"
                ]
                st.session_state.pop("pending_analysis", None)
                pending_batch_id = None
                st.success("Batch analysis complete!")
            except TimeoutError as e:
                st.warning(str(e))
            except AnalysisAuthenticationError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"Could not retrieve pending batch: {e}")

        # Analyse button (full width)
        credentials_ready = (
            (
                ("anthropic" not in required_cascade_providers
                 or bool(anthropic_api_key))
                and
                ("openai" not in required_cascade_providers
                 or bool(openai_api_key))
            )
            if is_review_cascade
            else bool(api_key)
        )
        analyse_disabled = (
            (not uploaded_file and not pasted_text)
            or len(selected_frameworks) == 0
            or not credentials_ready
            or bool(pending_batch_id)
        )

        if st.button(
            "Analyse Report", disabled=analyse_disabled, type="primary"
        ):
            if is_review_cascade and not credentials_ready:
                missing_providers = []
                if (
                    "anthropic" in required_cascade_providers
                    and not anthropic_api_key
                ):
                    missing_providers.append("Anthropic")
                if (
                    "openai" in required_cascade_providers
                    and not openai_api_key
                ):
                    missing_providers.append("OpenAI")
                st.error(
                    "Please enter the required "
                    + " and ".join(missing_providers)
                    + " API key(s)"
                )
            elif not is_review_cascade and not api_key:
                provider = get_model_config(selected_model_id)["provider"].title()
                st.error(f"Please enter your {provider} API key")
            elif len(selected_frameworks) == 0:
                st.error("Please select at least one framework")
            elif not uploaded_file and not pasted_text:
                st.error("Please upload a PDF or paste text")
            else:
                for state_key in (
                    "analysis_results",
                    "framework_summaries",
                    "num_pages",
                    "token_usage",
                    "analysis_strategy_used",
                ):
                    st.session_state.pop(state_key, None)
                if uploaded_file:
                    if page_end is not None and page_start > page_end:
                        st.error("'From page' must be \u2264 'To page'.")
                        st.stop()
                    with st.spinner("Extracting text from PDF..."):
                        try:
                            report_pages = extract_pdf_pages(
                                uploaded_file,
                                first_page=page_start,
                                last_page=page_end,
                                include_vision=use_vision,
                                max_vision_pages=30,
                            )
                            total = total_pages
                            end_idx = page_end if page_end is not None else total
                            vision_page_count = sum(
                                bool(page.get("image_base64"))
                                for page in report_pages
                            )
                            st.success(
                                f"Analysing pages {page_start}\u2013{end_idx} "
                                f"({len(report_pages)} of {total} pages)"
                            )
                            if use_vision:
                                st.info(
                                    f"Vision enabled for {vision_page_count} "
                                    "visually dense/scanned pages."
                                )
                        except Exception as e:
                            st.error(f"Failed to extract PDF: {e}")
                            st.stop()
                else:
                    report_pages = [
                        {"page_number": i, "text": paragraph.strip()}
                        for i, paragraph in enumerate(
                            (p for p in pasted_text.split('\n\n') if p.strip()),
                            start=1,
                        )
                    ]
                    st.info(f"Processing {len(report_pages)} text sections")

                report_text = format_report_text(report_pages)

                selected_model = get_model_config(selected_model_id)
                if is_review_cascade:
                    st.markdown(
                        "### Running reviewed cascade: "
                        f"{analyst_label} → {reviewer_label} → conditional "
                        f"{senior_label}..."
                    )
                else:
                    st.markdown(
                        f"### Analysing with {selected_model['label']}..."
                    )
                progress_bar = st.progress(0)

                if use_batch_api and not is_review_cascade:
                    st.session_state.pending_analysis = {
                        "batch_id": None,
                        "provider": selected_model["provider"],
                        "model_id": selected_model_id,
                        "report_text": report_text,
                        "report_pages": report_pages,
                        "selected_frameworks": list(selected_frameworks),
                        "num_pages": len(report_pages),
                    }

                try:
                    if is_review_cascade:
                        results, framework_summaries, token_usage = (
                            run_review_cascade(
                                report_text,
                                selected_frameworks,
                                anthropic_api_key,
                                openai_api_key,
                                framework_requirements,
                                progress_bar,
                                requirement_refs,
                                report_pages=report_pages,
                                analyst_model_id=analyst_model_id,
                                reviewer_model_id=reviewer_model_id,
                                senior_reviewer_model_id=(
                                    senior_reviewer_model_id
                                ),
                            )
                        )
                    else:
                        results, framework_summaries, token_usage = (
                            run_model_analysis(
                                report_text, selected_frameworks,
                                api_key, framework_requirements, progress_bar,
                                requirement_refs,
                                report_pages=report_pages,
                                use_batch=use_batch_api,
                                track_pending_batch=use_batch_api,
                                model_id=selected_model_id,
                            )
                        )
                    st.session_state.analysis_results = results
                    st.session_state.framework_summaries = (
                        framework_summaries
                    )
                    st.session_state.num_pages = len(report_pages)
                    st.session_state.token_usage = token_usage
                    st.session_state.analysis_strategy_used = (
                        analysis_strategy
                    )
                    st.session_state.pop("pending_analysis", None)
                    if (
                        is_review_cascade
                        and not token_usage.get("cascade_complete", True)
                    ):
                        failed_stages = token_usage.get(
                            "cascade_failure_stages"
                        ) or [
                            token_usage.get(
                                "cascade_failure_stage", "review"
                            )
                        ]
                        failed_stage_labels = [
                            str(stage).replace("_", " ")
                            for stage in failed_stages
                        ]
                        failed_stage_subject = (
                            f"The {failed_stage_labels[0]} stage"
                            if len(failed_stage_labels) == 1
                            else (
                                "The "
                                + " and ".join(failed_stage_labels)
                                + " stages"
                            )
                        )
                        failure_details = token_usage.get(
                            "cascade_failure_details", {}
                        )
                        detail_messages = [
                            str(detail.get("message", "")).strip()
                            for detail in failure_details.values()
                            if str(detail.get("message", "")).strip()
                        ]
                        failure_reason = (
                            f" {' '.join(dict.fromkeys(detail_messages))}"
                            if detail_messages
                            else ""
                        )
                        st.warning(
                            f"{failed_stage_subject} did not "
                            "complete for every requirement. "
                            "A bounded missing-only retry was attempted. "
                            "Earlier successful results were retained, and "
                            "remaining missing items are provisional for human "
                            "review."
                            f"{failure_reason}"
                        )
                    else:
                        st.success("Analysis complete!")
                except TimeoutError as e:
                    st.warning(str(e))
                except AnalysisAuthenticationError as e:
                    st.session_state.pop("pending_analysis", None)
                    st.error(str(e))
                except Exception as e:
                    st.error(f"Analysis failed: {e}")

        # --- Results (full width, below controls) ---
        st.markdown("---")

        if (
            'analysis_results' in st.session_state
            and st.session_state.analysis_results
        ):
            results = st.session_state.analysis_results
            framework_summaries = st.session_state.framework_summaries
            num_pages = st.session_state.num_pages
            token_usage = st.session_state.get('token_usage', {})

            total_results = len(results)
            scored_results = [
                result
                for result in results
                if not is_provisional_cascade_result(result)
            ]
            provisional_count = total_results - len(scored_results)
            covers_count = sum(
                1 for r in scored_results
                if r['classification'] == CLASSIFICATION_COVERS
            )
            partly_count = sum(
                1 for r in scored_results
                if r['classification'] == CLASSIFICATION_PARTLY
            )
            doesnt_count = sum(
                1 for r in scored_results
                if r['classification'] == CLASSIFICATION_DOESNT
            )
            scored_frameworks = {
                framework: summary
                for framework, summary in framework_summaries.items()
                if summary.get(
                    "scored_total",
                    summary.get("total", 0),
                )
            }
            best_fw = (
                max(
                    scored_frameworks.items(),
                    key=lambda x: x[1]['avg_score']
                )
                if scored_frameworks else None
            )

            # ── Coverage summary (donut + counts) ──
            overall_pct = (
                sum(
                    classification_to_score(r["classification"])
                    for r in scored_results
                ) / len(scored_results) * 100
            ) if scored_results else 0.0
            overall_label = f"{overall_pct:.0f}%" if scored_results else "N/A"
            best_note = (
                f'Best alignment with <strong class="t-strong">'
                f'{best_fw[0]}</strong>.'
                if best_fw else ''
            )
            donut_html = (
                f'<div style="display:grid;grid-template-columns:auto 1fr 1fr 1fr;'
                f'gap:14px;margin-bottom:16px;align-items:stretch;">'
                f'<div class="terra-dark-card" style="background:#0F3D2A;border-radius:14px;'
                f'padding:22px 26px;display:flex;align-items:center;gap:20px;'
                f'min-width:330px;">'
                f'<div style="width:104px;height:104px;border-radius:50%;'
                f'background:conic-gradient(#7FCB9E 0 {overall_pct:.1f}%, '
                f'rgba(255,255,255,.16) {overall_pct:.1f}% 100%);display:flex;'
                f'align-items:center;justify-content:center;flex-shrink:0;">'
                f'<div style="width:76px;height:76px;border-radius:50%;'
                f'background:#0F3D2A;display:flex;flex-direction:column;'
                f'align-items:center;justify-content:center;">'
                f'<span class="t-strong" style="font-family:\'Spectral\',serif;'
                f'font-size:26px;font-weight:600;color:#FCFAF3;line-height:1;">'
                f'{overall_label}</span>'
                f'<span class="t-soft" style="font-size:9px;color:#9FBAA8;'
                f'letter-spacing:.05em;">OVERALL</span>'
                f'</div></div>'
                f'<div><p class="t-soft" style="margin:0 0 4px;font-size:13px;'
                f'color:#9FBAA8;">'
                f'Resolved coverage</p>'
                f'<p class="t-strong" style="margin:0;font-size:13.5px;'
                f'line-height:1.5;color:#FCFAF3;">Analysed '
                f'<strong class="t-strong">{num_pages}</strong> pages '
                f'against <strong class="t-strong">{len(framework_summaries)}</strong> frameworks '
                f'({total_results} requirements'
                f'{f", {provisional_count} provisional and excluded from the score" if provisional_count else ""}). '
                f'{best_note}</p>'
                f'</div></div>'
                f'<div style="background:#E8F2EA;border:1px solid #C6E0CC;'
                f'border-radius:14px;padding:18px;">'
                f'<div style="font-family:\'Spectral\',serif;font-size:34px;'
                f'font-weight:600;color:#1C6B4A;line-height:1;">'
                f'{covers_count}</div>'
                f'<div style="font-weight:600;font-size:13px;color:#1C6B4A;'
                f'margin-top:5px;">Covers</div>'
                f'<div style="font-size:12px;color:#5B7A64;">fully addressed'
                f'</div></div>'
                f'<div style="background:#FBF0D8;border:1px solid #EBD6A3;'
                f'border-radius:14px;padding:18px;">'
                f'<div style="font-family:\'Spectral\',serif;font-size:34px;'
                f'font-weight:600;color:#B07A18;line-height:1;">'
                f'{partly_count}</div>'
                f'<div style="font-weight:600;font-size:13px;color:#B07A18;'
                f'margin-top:5px;">Partly covers</div>'
                f'<div style="font-size:12px;color:#977322;">needs strengthening'
                f'</div></div>'
                f'<div style="background:#F8E3DD;border:1px solid #EEC2B4;'
                f'border-radius:14px;padding:18px;">'
                f'<div style="font-family:\'Spectral\',serif;font-size:34px;'
                f'font-weight:600;color:#B4472F;line-height:1;">'
                f'{doesnt_count}</div>'
                f'<div style="font-weight:600;font-size:13px;color:#B4472F;'
                f'margin-top:5px;">Doesn\'t cover</div>'
                f'<div style="font-size:12px;color:#96543F;">action required'
                f'</div></div>'
                f'</div>'
            )
            st.markdown(donut_html, unsafe_allow_html=True)

            # ── Coverage by framework (stacked bars) ──
            bars_rows = ""
            for fw, s in sorted(
                framework_summaries.items(),
                key=lambda x: -x[1]["avg_score"]
            ):
                cts = s.get("counts", {})
                scored_total = int(
                    s.get("scored_total", s.get("total", 0)) or 0
                )
                if scored_total:
                    c_pct = (
                        cts.get(CLASSIFICATION_COVERS, 0)
                        / scored_total
                        * 100
                    )
                    p_pct = (
                        cts.get(CLASSIFICATION_PARTLY, 0)
                        / scored_total
                        * 100
                    )
                    d_pct = max(0.0, 100.0 - c_pct - p_pct)
                else:
                    c_pct = p_pct = d_pct = 0.0
                fw_pct = s.get("avg_score", 0) * 100
                fw_score_label = (
                    f"{fw_pct:.0f}%" if scored_total else "N/A"
                )
                fw_provisional = int(s.get("provisional", 0) or 0)
                provisional_note = (
                    f" &middot; {fw_provisional} provisional"
                    if fw_provisional
                    else ""
                )
                bars_rows += (
                    f'<div style="margin-bottom:14px;">'
                    f'<div style="display:flex;justify-content:space-between;'
                    f'margin-bottom:5px;">'
                    f'<span style="font-weight:600;font-size:13.5px;'
                    f'color:#152018;">{fw}</span>'
                    f'<span style="font-family:\'IBM Plex Mono\',monospace;'
                    f'font-size:11.5px;color:#8A9488;">{fw_score_label} &middot; '
                    f'{s.get("total", 0)} reqs'
                    f'{provisional_note}'
                    f'</span></div>'
                    f'<div style="height:11px;border-radius:6px;overflow:hidden;'
                    f'display:flex;background:#DDD5C2;">'
                    f'<div style="width:{c_pct:.1f}%;background:#1C6B4A;"></div>'
                    f'<div style="width:{p_pct:.1f}%;background:#C98A2B;"></div>'
                    f'<div style="width:{d_pct:.1f}%;background:#B4472F;"></div>'
                    f'</div></div>'
                )
            legend_html = (
                '<div style="display:flex;gap:18px;margin-top:16px;'
                'padding-top:14px;border-top:1px solid #E9E3D3;">'
                '<span style="display:inline-flex;align-items:center;gap:7px;'
                'font-size:12px;color:#4B5A50;"><span style="width:11px;'
                'height:11px;border-radius:3px;background:#1C6B4A;'
                'display:inline-block;"></span>Covers</span>'
                '<span style="display:inline-flex;align-items:center;gap:7px;'
                'font-size:12px;color:#4B5A50;"><span style="width:11px;'
                'height:11px;border-radius:3px;background:#C98A2B;'
                'display:inline-block;"></span>Partly</span>'
                '<span style="display:inline-flex;align-items:center;gap:7px;'
                'font-size:12px;color:#4B5A50;"><span style="width:11px;'
                'height:11px;border-radius:3px;background:#B4472F;'
                'display:inline-block;"></span>Doesn\'t</span>'
                '</div>'
            )
            st.markdown(
                f'<div style="background:#FCFAF3;border:1px solid #DDD5C2;'
                f'border-radius:14px;padding:20px 22px;margin-bottom:16px;">'
                f'<p style="margin:0 0 14px;font-weight:700;font-size:15px;'
                f'color:#152018;">Coverage by framework</p>'
                f'{bars_rows}{legend_html}</div>',
                unsafe_allow_html=True
            )

            cascade_results = [
                result for result in results
                if is_review_cascade_result(result)
            ]
            if cascade_results:
                cascade_counts = {
                    status: sum(
                        result.get("cascade_status") == status
                        for result in cascade_results
                    )
                    for status in CASCADE_STATUS_LABELS
                }
                cascade_human_review = sum(
                    result_needs_human_review(result)
                    for result in cascade_results
                )
                incomplete_reviews = sum(
                    cascade_counts.get(status, 0)
                    for status in (
                        "reviewer_failed",
                        "senior_reviewer_failed",
                        "luna_review_failed",
                        "terra_review_failed",
                    )
                )
                agreement_count = (
                    cascade_counts.get("analyst_reviewer_agree", 0)
                    + cascade_counts.get("haiku_luna_agree", 0)
                )
                adjudication_count = (
                    cascade_counts.get("senior_reviewer_adjudicated", 0)
                    + cascade_counts.get("terra_adjudicated", 0)
                )
                cascade_example = cascade_results[0]
                analyst_metric_label = cascade_role_label(
                    cascade_example, "analyst"
                )
                reviewer_metric_label = cascade_role_label(
                    cascade_example, "reviewer"
                )
                senior_metric_label = cascade_role_label(
                    cascade_example, "senior_reviewer"
                )
                st.markdown("### Reviewed cascade checks")
                metric_columns = st.columns(5)
                metric_columns[0].metric(
                    f"{analyst_metric_label} + {reviewer_metric_label} agree",
                    agreement_count,
                )
                metric_columns[1].metric(
                    f"{senior_metric_label} adjudications",
                    adjudication_count,
                )
                metric_columns[2].metric(
                    "Three-way disagreements",
                    cascade_counts["three_way_disagreement"],
                )
                metric_columns[3].metric(
                    "Incomplete reviews",
                    incomplete_reviews,
                )
                metric_columns[4].metric(
                    "Human review",
                    cascade_human_review,
                )
                if provisional_count:
                    st.warning(
                        f"{provisional_count} provisional cascade verdict(s) "
                        "are excluded from coverage percentages until a human "
                        "confirms them."
                    )

            # Cost estimate
            if token_usage:
                models_used = token_usage.get('models_used', set())
                usage_records = token_usage.get("usage_records", [])
                try:
                    total_cost, cache_savings = estimate_usage_cost(
                        usage_records
                    )
                    model_label = " + ".join(
                        get_model_config(model_id)["label"]
                        for model_id in sorted(models_used)
                    ) or get_model_config(
                        token_usage.get("selected_model", PRIMARY_MODEL)
                    )["label"]
                    if cascade_results:
                        model_label = f"Reviewed cascade ({model_label})"
                except ValueError as error:
                    st.warning(
                        f"Could not estimate this saved run's cost: {error}"
                    )
                else:
                    itok = (
                        token_usage.get("input_tokens", 0)
                        + token_usage.get("cache_read_tokens", 0)
                        + token_usage.get("cache_write_tokens", 0)
                    )
                    otok = token_usage.get("output_tokens", 0)
                    cache_str = (
                        f" \u00b7 Cache saved ~${cache_savings:.4f}"
                        if cache_savings > 0 else ""
                    )
                    batch_str = (
                        " \u00b7 Batch API 50% pricing applied"
                        if any(
                            record.get("batch_priced")
                            for record in usage_records
                        ) else ""
                    )
                    st.markdown(
                        f'<div style="background:#EDE7D8;border:1px solid '
                        f'#DDD5C2;border-radius:8px;padding:12px;'
                        f'margin-bottom:16px;font-size:13px;color:#3B4A40;">'
                        f'<strong>Model:</strong> {model_label} \u00b7 '
                        f'<strong>Estimated cost:</strong> ${total_cost:.4f} '
                        f'({itok:,} input / {otok:,} output tokens)'
                        f'{cache_str}{batch_str}'
                        f'</div>',
                        unsafe_allow_html=True
                    )

            # Highest-value review queue: uncertain verdicts are surfaced
            # before the full framework-by-framework result set.
            human_review_results = []
            seen_review_keys = set()
            for result in results:
                if not result_needs_human_review(result):
                    continue
                review_key = (
                    result.get("framework", ""),
                    result.get("requirement_id", ""),
                    result.get("reference", ""),
                    result.get("requirement", ""),
                )
                if review_key in seen_review_keys:
                    continue
                seen_review_keys.add(review_key)
                human_review_results.append(result)

            if human_review_results:
                st.markdown("### Human review queue")
                st.warning(
                    f"{len(human_review_results)} verdict(s) need review "
                    "first because of low confidence, model disagreement, "
                    "or both."
                )
                with st.expander(
                    f"Review {len(human_review_results)} uncertain verdict(s)",
                    expanded=True,
                ):
                    for result in human_review_results:
                        review_details = []
                        if result.get("needs_human_review"):
                            review_details.append(
                                cascade_status_label(result)
                                or "Cascade marked this for review"
                            )
                        if result.get("confidence", "low") == "low":
                            review_details.append(
                                result.get("confidence_reason", "")
                                or "The model did not provide a clear "
                                "confidence reason."
                            )
                        st.markdown(
                            f"**{result['framework']} · "
                            f"{prettify_topic_name(result['topic'])} — "
                            f"{result['classification']}**  \n"
                            f"{result.get('requirement', '')}  \n"
                            f"*Why review:* {'; '.join(review_details)}"
                        )
                        _, audit_trail = build_cascade_review_html(result)
                        if audit_trail:
                            st.markdown(
                                audit_trail,
                                unsafe_allow_html=True,
                            )
                        else:
                            extracts = result.get("relevant_extracts", [])
                            if isinstance(extracts, list) and extracts:
                                st.markdown(
                                    "**Evidence:**  \n"
                                    + "  \n".join(
                                        f"- {extract}" for extract in extracts
                                    )
                                )
                        st.markdown("---")

            # Export button
            excel_data = generate_results_excel(results, framework_summaries)
            st.download_button(
                label="\U0001f4e5 Download Results as Excel",
                data=excel_data,
                file_name="framework_analysis_results.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
            )

            # == DETAILED ANALYSIS (first) ==
            st.markdown("---")
            st.subheader(
                "Detailed Analysis \u2014 How does your report perform "
                "against selected frameworks?"
            )

            analysed_frameworks = list(
                dict.fromkeys(
                    result["framework"]
                    for result in results
                    if result.get("framework")
                )
            )
            for framework in analysed_frameworks:
                fw_results = [
                    r for r in results if r['framework'] == framework
                ]
                if not fw_results:
                    continue

                summary = framework_summaries.get(framework, {})
                counts = summary.get("counts", {})
                c_count = counts.get(CLASSIFICATION_COVERS, 0)
                p_count = counts.get(CLASSIFICATION_PARTLY, 0)
                d_count = counts.get(CLASSIFICATION_DOESNT, 0)
                review_count = sum(
                    result_needs_human_review(result)
                    for result in fw_results
                )

                with st.expander(
                    f"**{framework}** \u2014 {c_count} covered \u00b7 "
                    f"{p_count} partly \u00b7 {d_count} not covered \u00b7 "
                    f"{review_count} review first",
                    expanded=True
                ):
                    topics_seen = []
                    for r in fw_results:
                        if r["topic"] not in topics_seen:
                            topics_seen.append(r["topic"])

                    for topic in topics_seen:
                        topic_results = [
                            r for r in fw_results if r["topic"] == topic
                        ]
                        confidence_order = {"low": 0, "medium": 1, "high": 2}
                        topic_results.sort(
                            key=lambda item: confidence_order.get(
                                item.get("confidence", "low"), 0
                            )
                        )
                        st.markdown(f"**{prettify_topic_name(topic)}**")

                        for r in topic_results:
                            classification = r['classification']
                            clr = CLASSIFICATION_COLORS.get(
                                classification, "#8A9488"
                            )
                            badge_class = CLASSIFICATION_BADGES.get(
                                classification, "badge-doesnt"
                            )

                            extracts = r.get("relevant_extracts", [])
                            if extracts:
                                extracts_html = "".join(
                                    f'<div style="background:#F5F1E8;'
                                    f'border-left:3px solid {clr};'
                                    f'padding:6px 10px;margin:4px 0;'
                                    f'border-radius:0 4px 4px 0;'
                                    f'font-size:12px;color:#4B5A50;'
                                    f'font-style:italic;">'
                                    f'"{html.escape(str(ext))}"</div>'
                                    for ext in extracts
                                )
                                extracts_section = (
                                    '<p style="margin:8px 0 4px 0;'
                                    'font-size:11px;color:#8A9488;'
                                    'text-transform:uppercase;'
                                    'letter-spacing:0.5px;">'
                                    'Relevant text found:</p>'
                                    f'{extracts_html}'
                                )
                            else:
                                extracts_section = (
                                    '<p style="margin:8px 0 4px 0;'
                                    'font-size:12px;color:#B4472F;'
                                    'font-style:italic;">'
                                    'No relevant text found in report</p>'
                                )

                            req_text = r.get("requirement", "")
                            if len(req_text) > 200:
                                req_text = req_text[:200] + "\u2026"
                            req_text = html.escape(str(req_text))

                            ref = r.get("reference", "")
                            ref_html = (
                                f'<span style="font-size:11px;color:#8A9488;'
                                f'font-family:monospace;background:#DDD5C2;'
                                f'padding:1px 6px;border-radius:4px;'
                                f'margin-right:6px;white-space:nowrap;">'
                                f'{html.escape(str(ref))}</span>'
                                if ref else ""
                            )
                            confidence = r.get("confidence", "low")
                            confidence_colors = {
                                "low": ("#F8E3DD", "#B4472F"),
                                "medium": ("#FBF0D8", "#977322"),
                                "high": ("#E8F2EA", "#1C6B4A"),
                            }
                            confidence_bg, confidence_fg = confidence_colors.get(
                                confidence, confidence_colors["low"]
                            )
                            confidence_html = (
                                f'<span style="white-space:nowrap;background:'
                                f'{confidence_bg};color:{confidence_fg};padding:'
                                f'3px 8px;border-radius:10px;font-size:10px;'
                                f'font-weight:700;text-transform:uppercase;">'
                                f'{html.escape(str(confidence))} confidence</span>'
                            )
                            (
                                cascade_badge_html,
                                cascade_review_html,
                            ) = build_cascade_review_html(r)

                            st.markdown(
                                f'<div style="background:#EDE7D8;'
                                f'padding:12px;border-radius:8px;'
                                f'margin:8px 0;border-left:4px solid '
                                f'{clr};">'
                                f'<div style="display:flex;'
                                f'justify-content:space-between;'
                                f'align-items:flex-start;gap:12px;'
                                f'flex-wrap:wrap;">'
                                f'<span style="font-size:13px;'
                                f'color:#152018;flex:1;">'
                                f'{ref_html}{req_text}</span>'
                                f'<span class="{badge_class}" '
                                f'style="white-space:nowrap;">'
                                f'{html.escape(str(classification))}</span>'
                                f'{confidence_html}'
                                f'{cascade_badge_html}'
                                f'</div>'
                                f'{extracts_section}'
                                f'<p style="margin:8px 0 0 0;'
                                f'font-size:12px;color:#152018;">'
                                f'<strong>Rationale:</strong> '
                                f'{html.escape(str(r.get("rationale") or ""))}</p>'
                                f'<p style="margin:5px 0 0 0;font-size:11px;'
                                f'color:#4B5A50;"><strong>Confidence:</strong> '
                                f'{html.escape(str(r.get("confidence_reason") or ""))}</p>'
                                f'{cascade_review_html}'
                                f'</div>',
                                unsafe_allow_html=True
                            )

            # == GAP ANALYSIS (second) ==
            st.markdown("---")
            render_gap_analysis(results, framework_summaries)

        else:
            st.info(
                "Upload a document and click 'Analyse Report' "
                "to see results"
            )

    # ============================================
    # TAB 3: SIDE-BY-SIDE COMPARISON
    # ============================================
    with tab3:
        st.header("Side-by-Side Report Comparison")
        st.markdown(
            "Upload two ESG reports to compare how each covers the same "
            "framework requirements. Useful for benchmarking year-on-year "
            "progress or comparing two firms' disclosures."
        )

        cmp_model_id = st.selectbox(
            "Comparison model",
            options=USER_SELECTABLE_MODELS,
            index=0,
            format_func=model_picker_label,
            key="cmp_model_id",
            help="The same model is used for both reports.",
        )
        render_model_price_caption(cmp_model_id)
        cmp_provider = get_model_config(cmp_model_id)["provider"]
        cmp_api_key = render_model_api_key(
            cmp_model_id, f"cmp_{cmp_provider}_api_key"
        )
        st.caption(
            "Comparisons use standard API requests so two independent pending "
            "batches cannot be stranded. Estimated costs are shown afterwards."
        )

        available_frameworks = (
            list(framework_requirements.keys())
            if framework_requirements
            else list(FRAMEWORK_COLORS.keys())
        )

        st.subheader("Select Frameworks to Compare")

        cmp_col_sel1, cmp_col_sel2 = st.columns(2)
        with cmp_col_sel1:
            if st.button("Select All", key="cmp_select_all"):
                st.session_state.cmp_selected_frameworks = (
                    available_frameworks.copy()
                )
        with cmp_col_sel2:
            if st.button("Clear All", key="cmp_clear_all"):
                st.session_state.cmp_selected_frameworks = []

        if 'cmp_selected_frameworks' not in st.session_state:
            st.session_state.cmp_selected_frameworks = ["TCFD", "TNFD"]

        cmp_selected = []
        cmp_cols = st.columns(3)
        for i, fw in enumerate(available_frameworks):
            with cmp_cols[i % 3]:
                req_count = sum(
                    len(reqs)
                    for reqs in framework_requirements.get(fw, {}).values()
                )
                checked = st.checkbox(
                    f"{fw}",
                    value=fw in st.session_state.cmp_selected_frameworks,
                    key=f"cmp_fw_{fw}",
                    help=(
                        f"{FRAMEWORK_FULL_NAMES.get(fw, fw)} "
                        f"({req_count} requirements)"
                    )
                )
                if checked:
                    cmp_selected.append(fw)
        st.session_state.cmp_selected_frameworks = cmp_selected

        cmp_total_reqs = sum(
            sum(
                len(reqs)
                for reqs in framework_requirements.get(fw, {}).values()
            )
            for fw in cmp_selected
        )
        st.markdown(
            f"**{len(cmp_selected)}** framework(s) \u00b7 "
            f"**{cmp_total_reqs}** requirements each"
        )

        # Two-column upload
        st.subheader("Upload Two Reports")
        up_col1, up_col2 = st.columns(2)

        with up_col1:
            st.markdown("**Report A**")
            cmp_name_a = st.text_input(
                "Label for Report A", value="Report A", key="cmp_name_a"
            )
            cmp_file_a = st.file_uploader(
                "Upload PDF", type="pdf", key="cmp_file_a"
            )
            cmp_page_start_a, cmp_page_end_a = 1, None
            if cmp_file_a:
                import pymupdf
                bytes_a = cmp_file_a.read()
                cmp_file_a.seek(0)
                with pymupdf.open(stream=bytes_a, filetype="pdf") as doc:
                    total_a = len(doc)
                st.markdown(f"*{total_a} pages*")
                pa_c1, pa_c2 = st.columns(2)
                with pa_c1:
                    cmp_page_start_a = st.number_input(
                        "From page", 1, total_a, 1, key="cmp_ps_a"
                    )
                with pa_c2:
                    cmp_page_end_a = st.number_input(
                        "To page", 1, total_a, total_a, key="cmp_pe_a"
                    )

        with up_col2:
            st.markdown("**Report B**")
            cmp_name_b = st.text_input(
                "Label for Report B", value="Report B", key="cmp_name_b"
            )
            cmp_file_b = st.file_uploader(
                "Upload PDF", type="pdf", key="cmp_file_b"
            )
            cmp_page_start_b, cmp_page_end_b = 1, None
            if cmp_file_b:
                import pymupdf
                bytes_b = cmp_file_b.read()
                cmp_file_b.seek(0)
                with pymupdf.open(stream=bytes_b, filetype="pdf") as doc:
                    total_b = len(doc)
                st.markdown(f"*{total_b} pages*")
                pb_c1, pb_c2 = st.columns(2)
                with pb_c1:
                    cmp_page_start_b = st.number_input(
                        "From page", 1, total_b, 1, key="cmp_ps_b"
                    )
                with pb_c2:
                    cmp_page_end_b = st.number_input(
                        "To page", 1, total_b, total_b, key="cmp_pe_b"
                    )

        cmp_disabled = (
            (not cmp_file_a or not cmp_file_b)
            or len(cmp_selected) == 0
            or not cmp_api_key
        )
        if st.button(
            "Compare Reports", disabled=cmp_disabled,
            type="primary", key="cmp_run"
        ):
            if not cmp_api_key:
                provider = get_model_config(cmp_model_id)["provider"].title()
                st.error(f"Please enter your {provider} API key")
            elif not cmp_file_a or not cmp_file_b:
                st.error("Please upload both PDFs")
            elif len(cmp_selected) == 0:
                st.error("Please select at least one framework")
            else:
                with st.spinner(f"Extracting text from {cmp_name_a}..."):
                    pages_a = extract_pdf_pages(
                        cmp_file_a,
                        first_page=cmp_page_start_a,
                        last_page=cmp_page_end_a,
                        include_vision=True,
                        max_vision_pages=20,
                    )

                with st.spinner(f"Extracting text from {cmp_name_b}..."):
                    pages_b = extract_pdf_pages(
                        cmp_file_b,
                        first_page=cmp_page_start_b,
                        last_page=cmp_page_end_b,
                        include_vision=True,
                        max_vision_pages=20,
                    )

                report_a = format_report_text(pages_a)
                report_b = format_report_text(pages_b)

                st.markdown(f"### Analysing {cmp_name_a}...")
                progress_a = st.progress(0)
                try:
                    results_a, summaries_a, usage_a = (
                        run_model_analysis(
                            report_a, cmp_selected, cmp_api_key,
                            framework_requirements, progress_a,
                            requirement_refs,
                            report_pages=pages_a,
                            use_batch=False,
                            model_id=cmp_model_id,
                        )
                    )
                except AnalysisAuthenticationError as e:
                    st.error(str(e))
                    results_a, summaries_a = [], {}
                except Exception as e:
                    st.error(f"Failed on {cmp_name_a}: {e}")
                    results_a, summaries_a = [], {}

                st.markdown(f"### Analysing {cmp_name_b}...")
                progress_b = st.progress(0)
                try:
                    results_b, summaries_b, usage_b = (
                        run_model_analysis(
                            report_b, cmp_selected, cmp_api_key,
                            framework_requirements, progress_b,
                            requirement_refs,
                            report_pages=pages_b,
                            use_batch=False,
                            model_id=cmp_model_id,
                        )
                    )
                except AnalysisAuthenticationError as e:
                    st.error(str(e))
                    results_b, summaries_b = [], {}
                except Exception as e:
                    st.error(f"Failed on {cmp_name_b}: {e}")
                    results_b, summaries_b = [], {}

                if results_a and results_b:
                    st.session_state.cmp_results_a = results_a
                    st.session_state.cmp_results_b = results_b
                    st.session_state.cmp_summaries_a = summaries_a
                    st.session_state.cmp_summaries_b = summaries_b
                    st.session_state.cmp_stored_name_a = cmp_name_a
                    st.session_state.cmp_stored_name_b = cmp_name_b
                    st.session_state.cmp_frameworks = cmp_selected
                    st.session_state.cmp_usage_a = usage_a
                    st.session_state.cmp_usage_b = usage_b
                    st.session_state.cmp_model_id_used = cmp_model_id
                    st.success("Comparison complete!")

        # --- Display comparison results ---
        if (
            'cmp_results_a' in st.session_state
            and st.session_state.cmp_results_a
        ):
            results_a = st.session_state.cmp_results_a
            results_b = st.session_state.cmp_results_b
            summaries_a = st.session_state.cmp_summaries_a
            summaries_b = st.session_state.cmp_summaries_b
            name_a = st.session_state.cmp_stored_name_a
            name_b = st.session_state.cmp_stored_name_b
            common_frameworks = st.session_state.cmp_frameworks
            usage_a = st.session_state.get("cmp_usage_a", {})
            usage_b = st.session_state.get("cmp_usage_b", {})
            model_id_used = st.session_state.get(
                "cmp_model_id_used", PRIMARY_MODEL
            )

            st.markdown("---")
            st.subheader("Comparison Results")

            if usage_a and usage_b:
                cost_a, _ = estimate_usage_cost(
                    usage_a.get("usage_records", [])
                )
                cost_b, _ = estimate_usage_cost(
                    usage_b.get("usage_records", [])
                )
                model_label = get_model_config(model_id_used)["label"]
                st.info(
                    f"{model_label} estimated cost — {name_a}: "
                    f"${cost_a:.4f}; {name_b}: ${cost_b:.4f}; combined: "
                    f"${cost_a + cost_b:.4f}. Standard API pricing applied."
                )

            st.markdown(
                '<div style="background:#EDE7D8;border:1px solid #DDD5C2;'
                'border-radius:8px;padding:16px;margin-bottom:16px;">'
                '<h4 style="margin:0 0 12px 0;color:#152018;">'
                'Coverage Summary</h4>'
                '<table style="width:100%;border-collapse:collapse;'
                'font-size:13px;">'
                '<tr style="border-bottom:2px solid #DDD5C2;">'
                '<th style="text-align:left;padding:6px;color:#152018;">'
                'Framework</th>'
                f'<th style="text-align:center;padding:6px;color:#152018;"'
                f' colspan="3">{name_a}</th>'
                f'<th style="text-align:center;padding:6px;color:#152018;"'
                f' colspan="3">{name_b}</th>'
                '</tr>'
                '<tr style="border-bottom:1px solid #DDD5C2;'
                'font-size:11px;color:#8A9488;">'
                '<td></td>'
                '<td style="text-align:center;padding:4px;">Covered</td>'
                '<td style="text-align:center;padding:4px;">Partly</td>'
                '<td style="text-align:center;padding:4px;">Not</td>'
                '<td style="text-align:center;padding:4px;">Covered</td>'
                '<td style="text-align:center;padding:4px;">Partly</td>'
                '<td style="text-align:center;padding:4px;">Not</td>'
                '</tr>',
                unsafe_allow_html=True
            )

            table_rows = ""
            for fw in common_frameworks:
                sa = summaries_a.get(fw, {}).get("counts", {})
                sb = summaries_b.get(fw, {}).get("counts", {})
                table_rows += (
                    '<tr style="border-bottom:1px solid #E9E3D3;">'
                    f'<td style="padding:6px;font-weight:600;'
                    f'color:#152018;">{fw}</td>'
                    f'<td style="text-align:center;padding:6px;'
                    f'background:#E8F2EA;color:#1C6B4A;">'
                    f'{sa.get(CLASSIFICATION_COVERS, 0)}</td>'
                    f'<td style="text-align:center;padding:6px;'
                    f'background:#FBF0D8;color:#B07A18;">'
                    f'{sa.get(CLASSIFICATION_PARTLY, 0)}</td>'
                    f'<td style="text-align:center;padding:6px;'
                    f'background:#F8E3DD;color:#B4472F;">'
                    f'{sa.get(CLASSIFICATION_DOESNT, 0)}</td>'
                    f'<td style="text-align:center;padding:6px;'
                    f'background:#E8F2EA;color:#1C6B4A;">'
                    f'{sb.get(CLASSIFICATION_COVERS, 0)}</td>'
                    f'<td style="text-align:center;padding:6px;'
                    f'background:#FBF0D8;color:#B07A18;">'
                    f'{sb.get(CLASSIFICATION_PARTLY, 0)}</td>'
                    f'<td style="text-align:center;padding:6px;'
                    f'background:#F8E3DD;color:#B4472F;">'
                    f'{sb.get(CLASSIFICATION_DOESNT, 0)}</td>'
                    '</tr>'
                )
            st.markdown(
                f'{table_rows}</table></div>', unsafe_allow_html=True
            )

            st.subheader("Requirement-by-Requirement")

            b_lookup = {}
            for r in results_b:
                key = (r["framework"], r["topic"], r["requirement"][:100])
                b_lookup[key] = r

            for fw in common_frameworks:
                fw_results_a = [
                    r for r in results_a if r["framework"] == fw
                ]
                if not fw_results_a:
                    continue

                better_a, better_b, same = 0, 0, 0
                for r_a in fw_results_a:
                    key = (
                        r_a["framework"], r_a["topic"],
                        r_a["requirement"][:100]
                    )
                    r_b = b_lookup.get(key)
                    if r_b:
                        s_a = classification_to_score(r_a["classification"])
                        s_b = classification_to_score(r_b["classification"])
                        if s_a > s_b:
                            better_a += 1
                        elif s_b > s_a:
                            better_b += 1
                        else:
                            same += 1

                with st.expander(
                    f"**{fw}** \u2014 {name_a} leads on {better_a} \u00b7 "
                    f"{name_b} leads on {better_b} \u00b7 {same} same"
                ):
                    topics_seen = []
                    for r in fw_results_a:
                        if r["topic"] not in topics_seen:
                            topics_seen.append(r["topic"])

                    for topic in topics_seen:
                        topic_results = [
                            r for r in fw_results_a if r["topic"] == topic
                        ]
                        st.markdown(f"**{prettify_topic_name(topic)}**")

                        for r_a in topic_results:
                            key = (
                                r_a["framework"], r_a["topic"],
                                r_a["requirement"][:100]
                            )
                            r_b = b_lookup.get(key)

                            class_a = r_a["classification"]
                            class_b = (
                                r_b["classification"]
                                if r_b else CLASSIFICATION_DOESNT
                            )
                            badge_a = CLASSIFICATION_BADGES.get(
                                class_a, "badge-doesnt"
                            )
                            badge_b = CLASSIFICATION_BADGES.get(
                                class_b, "badge-doesnt"
                            )

                            req_text = r_a["requirement"]
                            if len(req_text) > 180:
                                req_text = req_text[:180] + "\u2026"

                            st.markdown(
                                f'<div style="background:#EDE7D8;'
                                f'padding:12px;border-radius:8px;'
                                f'margin:8px 0;">'
                                f'<p style="margin:0 0 8px 0;'
                                f'font-size:13px;color:#152018;">'
                                f'{req_text}</p>'
                                f'<div style="display:flex;gap:16px;'
                                f'align-items:center;flex-wrap:wrap;">'
                                f'<div style="flex:1;min-width:200px;">'
                                f'<span style="font-size:11px;color:#8A9488;'
                                f'text-transform:uppercase;">'
                                f'{name_a}</span><br>'
                                f'<span class="{badge_a}">'
                                f'{class_a}</span>'
                                f'</div>'
                                f'<div style="flex:1;min-width:200px;">'
                                f'<span style="font-size:11px;color:#8A9488;'
                                f'text-transform:uppercase;">'
                                f'{name_b}</span><br>'
                                f'<span class="{badge_b}">'
                                f'{class_b}</span>'
                                f'</div>'
                                f'</div>'
                                f'</div>',
                                unsafe_allow_html=True
                            )

            st.markdown("---")
            cmp_excel = generate_comparison_excel(
                results_a, results_b, name_a, name_b, common_frameworks
            )
            st.download_button(
                label="\U0001f4e5 Download Comparison as Excel",
                data=cmp_excel,
                file_name="framework_comparison.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                key="cmp_download"
            )

    # ============================================
    # TAB 4: REPORT DRAFTER (BETA)
    # ============================================
    with tab4:
        report_drafter.render_drafter_tab(
            framework_requirements, requirement_refs,
            FRAMEWORK_FULL_NAMES, extract_text_from_pdf,
            prettify_topic_name,
        )


if __name__ == "__main__":
    main()
